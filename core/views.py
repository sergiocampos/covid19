from django.utils.translation import gettext as _
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash, login, authenticate, logout
from django.contrib.auth.forms import PasswordChangeForm, UserCreationForm
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from datetime import datetime, timezone
from .models import RegistroCovid, Cnes, Status, Regulacao
import calendar
from time import gmtime, strftime
from django.core.paginator import Paginator
from django.urls import reverse
from urllib.parse import urlencode
import pandas as pds
from django.http import HttpResponse, HttpResponseRedirect
from io import BytesIO as IO
from openpyxl import Workbook
from core.forms import UserForm, UserProfileInfoForm
from django.db.models import Max
import base64
from datetime import timedelta

# Create your views here.
def index(request):
	return render(request,'index.html')


@login_required
def special(request):
	return HttpResponse("You are logged in !")

@login_required
def user_logout(request):
	logout(request)
	return HttpResponseRedirect(reverse('index'))

def register(request):
	registered = False
	if request.method == 'POST':
		user_form = UserForm(data=request.POST)
		profile_form = UserProfileInfoForm(data=request.POST)
		if user_form.is_valid() and profile_form.is_valid():
			user = user_form.save()
			user.set_password(user.password)
			user.save()
			profile = profile_form.save(commit=False)
			profile.user = user
			#if 'profile_pic' in request.FILES:
			#	print('found it')
			#	profile.profile_pic = request.FILES['profile_pic']
			profile.save()
			registered = True
		else:
			print(user_form.errors,profile_form.errors)
	else:
		user_form = UserForm()
		profile_form = UserProfileInfoForm()
	return render(request,'registration.html', {'user_form':user_form,
		'profile_form':profile_form, 'registered':registered})


def user_login(request):
	if request.method == 'POST':
		username = request.POST.get('username')
		password = request.POST.get('password')
		user = authenticate(username=username, password=password)
		if user:
			if user.is_active:
				login(request,user)
				return HttpResponseRedirect(reverse('index'))
			else:
				return HttpResponse("Your account was inactive.")
		else:
			print("Someone tried to login and failed.")
			print("They used username: {} and password: {}".format(username,password))
			return HttpResponse("Invalid login details given")
	else:
		return render(request, 'login.html', {})


#codigo a seguir é outro!
@login_required
def signup(request):
	if request.method == 'POST':
		form = UserCreationForm(request.POST)
		if form.is_valid():
			form.save()
			username = form.cleaned_data.get('username')
			raw_password = form.cleaned_data.get('password1')
			user = authenticate(username=username, password=raw_password)
			login(request, user)
			return redirect('/')
	else:
		form = UserCreationForm()
	return render(request, 'signup.html', {'form': form})


@login_required
def change_password(request):
	if request.method == 'POST':
		form = PasswordChangeForm(request.user, request.POST)
		if form.is_valid():
			user = form.save()
			update_session_auth_hash(request, user)
			messages.success(request, _('Your password was successfully updated!'))
			return redirect('/')
		else:
			messages.error(request, _('Please correct the error below.'))
	else:
		form = PasswordChangeForm(request.user)
	return render(request, 'change_password.html', {'form': form})


@login_required
def covid_list(request):
	registros = RegistroCovid.objects.all().order_by('data_notificacao').reverse()

	if registros:
		paginator = Paginator(registros, 10)
		page = request.GET.get('page')
		regs = paginator.get_page(page)
		return render(request, 'list.html', {'regs':regs})
	else:
		return render(request, 'list.html')


@login_required
def search_estabelecimento(request):
	data = datetime.now()
	formateDate = data.strftime("%d-%m-%Y")
	hora = data.strftime("%H:%M")

	municipios = Cnes.objects.order_by('MUNICIPIO').distinct('MUNICIPIO')

	return render(request, 'search_estabelecimento.html', {'formateDate': formateDate, 'hora': hora, 'municipios':municipios})


@login_required
def search_estabelecimento_set(request):
	data = datetime.now()
	formateDate = data.strftime("%d-%m-%Y")
	hora = data.strftime("%H:%M")

	municipios = Cnes.objects.order_by('MUNICIPIO').distinct('MUNICIPIO')

	nome_solicitante = request.POST.get('nome_solicitante')
	estabelecimento_outro = request.POST.get('desc_outro_estabelecimento')
	nome_paciente = request.POST.get('nome_paciente')
	
	
	idade_paciente_cap = request.POST.get('idade_paciente')
	if idade_paciente_cap != '':
		idade_paciente = int(idade_paciente_cap)
	else:
		idade_paciente = None
	
	recurso_que_precisa = request.POST.get('recurso_que_precisa')
	estado_origem = request.POST.get('estado_paciente')
	cidade_origem = request.POST.get('cidade_paciente')

	cidade = request.POST.get('municipio_solicitante')

	request.session['nome_solicitante'] = nome_solicitante
	request.session['idade_paciente'] = idade_paciente
	request.session['nome_paciente'] = nome_paciente
	request.session['municipio_solicitante'] = cidade

	#context = {'cidade':cidade, 'idade_paciente':idade_paciente}
	#print(idade_paciente)
	#print(cidade)
	#base_url = reverse('registro_covid')
	#idade_paciente =  urlencode({'idade_paciente': idade_paciente})
	#cidade = urlencode({'cidade':cidade})
	#url = '{}?{}${}'.format(base_url, idade_paciente, cidade)

	#print(idade_paciente)
	#print(cidade)

	return redirect('registro_covid')


@login_required
def registro_covid(request):
	data = datetime.now()
	formateDate = data.strftime("%d-%m-%Y")
	hora = data.strftime("%H:%M")

	municipios = Cnes.objects.order_by('MUNICIPIO').distinct('MUNICIPIO')

	#idade_paciente = request.GET.get('idade_paciente')
	#cidade = request.GET.get('municipio_solicitante')
	nome_solicitante = request.session['nome_solicitante']
	idade_paciente = request.session['idade_paciente']
	nome_paciente = request.session['nome_paciente']
	cidade = request.session['municipio_solicitante']

	estabelecimentos = Cnes.objects.filter(MUNICIPIO = cidade).order_by('NO_FANTASIA')

	return render(request, 'registro_covid.html', {'formateDate': formateDate, 
		'hora': hora, 'municipios':municipios, 'idade_paciente':idade_paciente, 
		'cidade':cidade, 'estabelecimentos':estabelecimentos, 'nome_paciente':nome_paciente, 
		'nome_solicitante':nome_solicitante})



@login_required
def registro_covid_set(request):
	
	responsavel_pelo_preenchimento = request.user

	data = datetime.now()
	data_notificacao = data.strftime("%d-%m-%Y")
	hora_notificacao = data.strftime("%H:%M")

	last_registro = RegistroCovid.objects.all().last()
	if not last_registro:
		last_codigo_registro_total = 5000
		last_codigo_registro_mensal = 0
	else:
		last_codigo_registro_total = last_registro.codigo_registro_total
		last_codigo_registro_mensal = last_registro.codigo_registro_mensal
	

	#Dia de hoje:
	dia = data.day
	#mês:
	if last_registro:
		mes_ultimo_registro=last_registro.data_notificacao.month
		ano_ultimo_registro=last_registro.data_notificacao.year
	else:
		mes_ultimo_registro=data.month
		ano_ultimo_registro=data.year
	mes_corrente=data.month
	
	#Ano corrente
	ano_corrente = data.year
	
	#Verifica se o mês mudou:
	if (ano_corrente > ano_ultimo_registro):
		if (mes_corrente < mes_ultimo_registro):
			last_codigo_registro_mensal = 1
			last_codigo_registro_total += 1
		elif (mes_corrente >= mes_ultimo_registro):
			last_codigo_registro_mensal = 1
			last_codigo_registro_total += 1
	elif (ano_corrente == ano_ultimo_registro):
		if (mes_corrente > mes_ultimo_registro):
			last_codigo_registro_mensal = 1
			last_codigo_registro_total += 1
		else:
			last_codigo_registro_total += 1
			last_codigo_registro_mensal += 1


	#último dia do mês:
	#ultimo_dia = calendar.monthrange(int(strftime("%Y", gmtime())), int(strftime("%m", gmtime())))[1]
	#if dia <= ultimo_dia:
	#	last_codigo_registro_total += 1
	#	last_codigo_registro_mensal += 1
	#else:
	#	last_codigo_registro_total += 1
	#	last_codigo_registro_mensal += 0

	last_codigo_registro_total_str = str(last_codigo_registro_total)
	last_codigo_registro_mensal_str = str(last_codigo_registro_mensal)


	num_registro = last_codigo_registro_total_str + last_codigo_registro_mensal_str

	
	nome_solicitante = request.POST.get('nome_solicitante')
	municipio_estabelecimento_solicitante = request.POST.get('municipio_solicitante')
	estabelecimento_solicitante = request.POST.get('estabelecimento_solicitante_new')
	estabelecimento_solicitante_outro = request.POST.get('desc_outro_estabelecimento')
	nome_paciente = request.POST.get('nome_paciente')
	
	idade_paciente_cap = request.POST.get('idade_paciente')
	if idade_paciente_cap != '':
		idade_paciente = int(idade_paciente_cap)
	else:
		idade_paciente = None
	
	sexo_paciente = request.POST.get('sexo_paciente')
	recurso_que_precisa = request.POST.get('recurso_que_precisa')
	estado_origem = request.POST.get('estado_paciente')
	cidade_origem = request.POST.get('cidade_paciente')
	telefone_retorno = request.POST.get('telefone_retorno')

	data_admissao_cap = request.POST.get('data_admissao')
	if data_admissao_cap == '' or data_admissao_cap == None:
		data_admissao = None
	else:
		data_admissao = data_admissao_cap


	frequencia_respiratoria_cap = request.POST.get('frequencia_respiratoria')
	if frequencia_respiratoria_cap != '':
		frequencia_respiratoria = int(frequencia_respiratoria_cap)
	else:
		frequencia_respiratoria = None

	saturacao_paciente_cap = request.POST.get('saturacao_paciente')
	if saturacao_paciente_cap != '':
		saturacao_paciente = int(saturacao_paciente_cap)
	else:
		saturacao_paciente = None

	ar_o2 = request.POST.get('ar_o2')
	
	frequencia_cardiaca_cap = request.POST.get('f_cardiaca_paciente')
	if frequencia_cardiaca_cap != '':
		frequencia_cardiaca = int(frequencia_cardiaca_cap)
	else:
		frequencia_cardiaca = None

	pa_part1 = request.POST.get('pa_part1')
	pa_part2 = request.POST.get('pa_part2')

	pa_admissao = pa_part1 +"x"+ pa_part2

	pa_part1_ = request.POST.get('pa_part1_')
	pa_part2_ = request.POST.get('pa_part2_')

	pa = pa_part1_ +"x"+ pa_part2_

	conciencia = request.POST.get('consciencia_paciente')
	
	temperatura_cap = request.POST.get('temperatura_paciente')
	if temperatura_cap != '':
		temperatura = float(temperatura_cap)
	else:
		temperatura = None

	observacao = request.POST.get('observacao_paciente')


	#implementar o algoritmo: codigo_registro = None
	codigo_registro_total = last_codigo_registro_total_str
	codigo_registro_mensal = last_codigo_registro_mensal_str


	codigo_registro_completo = last_codigo_registro_total_str + last_codigo_registro_mensal_str

	image_descricao_clinica_cap = request.FILES.get('file_descricao_clinica')
	if image_descricao_clinica_cap == '' or image_descricao_clinica_cap == None:
		image_descricao_clinica = None
	else:
		image_descricao_clinica = image_descricao_clinica_cap.file.read()


	image_laudo_tc_cap = request.FILES.get('file_tc_torax')
	if image_laudo_tc_cap == '' or image_laudo_tc_cap == None:
		image_laudo_tc = None
	else:
		image_laudo_tc = image_laudo_tc_cap.file.read()
	
	image_laudo_rx_cap = request.FILES.get('file_rx_torax')
	if image_laudo_rx_cap == '' or image_laudo_rx_cap == None:
		image_laudo_rx = None
	else:
		image_laudo_rx = image_laudo_rx_cap.file.read()


	descricao_clinica = request.POST.get('descricao_clinica')
	
	sindrome_gripal = request.POST.getlist('s_gripal')
	
	tempo_quadro_sintomatico_cap = request.POST.get('tempo_inicio_sintomas')
	if tempo_quadro_sintomatico_cap == '' or tempo_quadro_sintomatico_cap == None:
		tempo_quadro_sintomatico = None
	else:
		tempo_quadro_sintomatico = int(float(tempo_quadro_sintomatico_cap))

	exposicao_pessoa_infectada = request.POST.get('exposicao_epidemiologica_sim')
	parentesco = request.POST.get('grau_parentesco')
	comorbidades = request.POST.getlist('comorbidades')
	outras_comorbidades = request.POST.get('descricao_outras_comorbidades')
	outras_cardiopatias = request.POST.get('desc_outras_cardiopatias')
	idade_gestacional = request.POST.get('idade_gestacao')
	comorbidades_obstetricas = request.POST.get('comorbidade_obstetrica')
	gesta_para = request.POST.get('gestacao_para')
	medicamentos_uso_habitual = request.POST.getlist('medicamento_uso_habitual')
	medicamento_outros = request.POST.get('descricao_outros_medicamento')
	
	spo2_cap = request.POST.get('spo2')
	if spo2_cap == '' or spo2_cap == None:
		spo2 = None
	else:
		spo2 = int(float(spo2_cap))
		

	fr_irpm_cap = request.POST.get('fr_irpm')
	if fr_irpm_cap == '' or fr_irpm_cap == None:
		fr_irpm = None
	else:
		fr_irpm = int(float(fr_irpm_cap))
		

	ventilacao_tipo = request.POST.get('ventilacao')
	o2_suporte = request.POST.getlist('o2_suporte')
	
	dose_cn_cap = request.POST.get('dose_cn')
	if dose_cn_cap == '' or dose_cn_cap == None:
		dose_cn = None
	else:
		dose_cn = float(dose_cn_cap)

	dose_venturi_cap = request.POST.get('dose_venturi')
	if dose_venturi_cap == '' or dose_venturi_cap == None:
		dose_venturi = None
	else:
		dose_venturi = float(dose_venturi_cap)

	mascara_com_reservatorio_cap = request.POST.get('valor_mascara')
	if mascara_com_reservatorio_cap == '' or mascara_com_reservatorio_cap == None:
		mascara_com_reservatorio = None
	else:
		mascara_com_reservatorio = mascara_com_reservatorio_cap

	

	vmi = request.POST.getlist('vmi')
	
	vt_cap = request.POST.get('vt')
	if vt_cap == '' or vt_cap == None:
		vt = None
	else:
		vt = float(vt_cap)
		

	delta_pressure_cap = request.POST.get('delta_pressure')
	if delta_pressure_cap == '' or delta_pressure_cap == None:
		delta_pressure = None
	else:
		delta_pressure = float(delta_pressure_cap)
		

	pplato_cap = request.POST.get('pplato')
	if pplato_cap == '' or pplato_cap == None:
		pplato = None
	else:
		pplato = float(pplato_cap)
		

	raw_cap = request.POST.get('raw')
	if raw_cap == '' or raw_cap == None:
		raw = None
	else:
		raw = float(raw_cap)
		

	cest_cap = request.POST.get('cest')
	if cest_cap == '' or cest_cap == None :
		cest = None
	else:
		cest = float(cest_cap)
		

	sao2_cap = request.POST.get('sao2')
	if sao2_cap == '' or sao2_cap == None:
		sao2 = None
	else:
		sao2 = float(sao2_cap)

	pao2_cap = request.POST.get('pao2')
	if pao2_cap == '' or pao2_cap == None:
		pao2 = None
	else:
		pao2 = float(pao2_cap)

	fio2_cap = request.POST.get('fio2')
	if fio2_cap == '' or fio2_cap == None:
		fio2 = None
	else:
		fio2 = float(fio2_cap)

	paco2_cap = request.POST.get('paco2')
	if paco2_cap == '' or paco2_cap == None:
		paco2 = None
	else:
		paco2 = float(paco2_cap)
	
	fc_cap = request.POST.get('fc')
	if fc_cap == '' or fc_cap == None:
		fc = None
	else:
		fc = int(float(fc_cap))
		

	temperatura_axilar_cap = request.POST.get('temp_auxiliar')
	if temperatura_axilar_cap == '' or temperatura_axilar_cap == None :
		temperatura_axilar = None
	else:
		temperatura_axilar = float(temperatura_axilar_cap)
		

	droga_vasoativa = request.POST.getlist('droga_vasoativa')
	
	qtd_nora_cap = request.POST.get('qtd_nora')
	if qtd_nora_cap == '' or qtd_nora_cap == None:
		qtd_nora = None
	else:
		qtd_nora = float(qtd_nora_cap)

	qtd_adrenalina_cap = request.POST.get('qtd_adrenalina')
	if qtd_adrenalina_cap == '' or qtd_adrenalina_cap == None:
		qtd_adrenalina = None
	else:
		qtd_adrenalina = float(qtd_adrenalina_cap)

	qtd_vasopressina_cap = request.POST.get('qtd_vasopressina')
	if qtd_vasopressina_cap == '' or qtd_vasopressina_cap == None:
		qtd_vasopressina = None
	else:
		qtd_vasopressina = float(qtd_vasopressina_cap)

	qtd_dobutamina_cap = request.POST.get('qtd_dobutamina')
	if qtd_dobutamina_cap == '' or qtd_dobutamina_cap == None:
		qtd_dobutamina = None
	else:
		qtd_dobutamina = float(qtd_dobutamina_cap)

	arritmia = request.POST.getlist('arritmia')
	infeccao_bacteriana = request.POST.get('infeccao_bacteriana')
	foco = request.POST.getlist('foco')
	uso_antibioticoterapia = request.POST.get('uso_antibioticoterapia')
	pesquisa_teste_sars_cov_2 = request.POST.get('pesquisa_sars_cov2')
	igg_cap = request.POST.get('igg')
	#igg_bd = request.POST.get('igg_bd')
	if igg_cap == '' or igg_cap == None:
		#if igg_bd == '' or igg_bd == None:
		igg = None
		#else:
		#	igg = igg_bd
	else:
		igg = igg_cap

	#igm_do_bd = registro.igm
	igm_cap = request.POST.get('igm')

	#igm_bd_cap = request.POST.get('igm_bd')
	if igm_cap == '' or igm_cap == None:
		#if igm_bd_cap == '' or igm_bd_cap == None:
		igm = None
		#else:
		#	igm = igm_bd_cap
	else:
		igm = igm_cap

	
	data_igm_igg_cap = request.POST.get('data_igm_igg')
	#data_igm_igg_bd_cap = request.POST.get('data_igm_igg_bd')

	if data_igm_igg_cap == '' or data_igm_igg_cap == None:
		#if data_igm_igg_bd_cap == '' or data_igm_igg_bd_cap == None:
		data_igm_igg = None
		#else:
		#	data_igm_igg = datetime.strptime(data_igm_igg_bd_cap, '%d/%m/%Y').date()
	else:
		data_igm_igg = data_igm_igg_cap


	rt_pcr_sars_cov_2 = request.POST.get('pcr_sars_cov2')
	
	#data_da_coleta_bd_cap = request.POST.get('data_da_coleta_bd')
	data_coleta_cap = request.POST.get('data_da_coleta')
	if data_coleta_cap == '' or data_coleta_cap == None:
		#if data_da_coleta_bd_cap == '' or data_da_coleta_bd_cap == None:
		data_coleta = None
		#else:
		#	data_coleta = datetime.strptime(data_da_coleta_bd_cap, '%d/%m/%Y').date()
	else:
		data_coleta = data_coleta_cap
	
	em_uso_corticosteroide = request.POST.get('em_uso_corticosteroide')
	dose_corticosteroide = request.POST.get('dose_corticosteroide')
	
	data_inicio_corticosteroide_bd_cap = request.POST.get('data_inicio_corticosteroide_bd')
	data_inicio_corticosteroide_cap = request.POST.get('data_inicio_corticosteroide')
	if data_inicio_corticosteroide_cap == '' or data_inicio_corticosteroide_cap == None:
		if data_inicio_corticosteroide_bd_cap == '' or data_inicio_corticosteroide_bd_cap == None:
			data_inicio_corticosteroide = None
		else:
			data_inicio_corticosteroide = datetime.strptime(data_inicio_corticosteroide_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_corticosteroide = data_inicio_corticosteroide_cap
	

	em_uso_hidroxicloroquina = request.POST.get('em_uso_hidrocloroquina')
	
	data_inicio_hidroxicloroquina_cap = request.POST.get('data_inicio_hidroxicloroquina')
	data_inicio_hidroxicloroquina_bd_cap = request.POST.get('data_inicio_hidroxicloroquina_bd')


	if data_inicio_hidroxicloroquina_cap == '' or data_inicio_hidroxicloroquina_cap == None:
		if data_inicio_hidroxicloroquina_bd_cap == '' or data_inicio_hidroxicloroquina_bd_cap == None:
			data_inicio_hidroxicloroquina = None	
		else:
			data_inicio_hidroxicloroquina = datetime.strptime(data_inicio_hidroxicloroquina_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_hidroxicloroquina = data_inicio_hidroxicloroquina_cap


	em_uso_oseltamivir = request.POST.get('em_uso_oseltamivir')
	
	data_inicio_oseltamivir_cap = request.POST.get('data_inicio_oseltamivir')
	data_inicio_oseltamivir_bd_cap = request.POST.get('data_inicio_oseltamivir_bd')


	if data_inicio_oseltamivir_cap == '' or data_inicio_oseltamivir_cap == None:
		if data_inicio_oseltamivir_bd_cap == '' or data_inicio_oseltamivir_bd_cap == None:
			data_inicio_oseltamivir = None
		else:
			data_inicio_oseltamivir = datetime.strptime(data_inicio_oseltamivir_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_oseltamivir = data_inicio_oseltamivir_cap



	em_uso_heparina = request.POST.get('em_uso_heparina')

	data_inicio_heparina_bd_cap = request.POST.get('data_inicio_heparina_bd')
	data_inicio_heparina_cap = request.POST.get('data_inicio_heparina')

	if data_inicio_heparina_cap == '' or data_inicio_heparina_cap == None:
		if data_inicio_heparina_bd_cap == '' or data_inicio_heparina_bd_cap == None:
			data_inicio_heparina = None
		else:
			data_inicio_heparina = datetime.strptime(data_inicio_heparina_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_heparina = data_inicio_heparina_cap
		

	tipo_heparina = request.POST.get('tipo_heparina')
	dose_heparina = request.POST.get('dose_heparina')
	
	pps_cap = request.POST.get('pps_paciente')
	if pps_cap == '' or pps_cap == None:
		pps = None
	else:
		pps = float(pps_cap)
		

	escala_pontos_glasgow_cap = request.POST.get('escala_glasgow')
	if escala_pontos_glasgow_cap == '' or escala_pontos_glasgow_cap == None:
		escala_pontos_glasgow = None
	else:
		escala_pontos_glasgow = int(float(escala_pontos_glasgow_cap))
		

	bloqueador_neuromuscular = request.POST.getlist('desc_bloqueador_neuromuscular')
	midazolam_dose = request.POST.get('midazolam_dose')
	fentanil_dose = request.POST.get('fentanil_dose')
	propofol_dose = request.POST.get('propofol_dose')
	sedacao_continua = request.POST.getlist('desc_sedacao_continua')
	rocuronio_dose = request.POST.get('rocuronio_dose')
	pacuronio_dose = request.POST.get('pacuronio_dose')
	cisatracurio_dose = request.POST.get('cisatracurio_dose')
	rass = request.POST.get('rass')
	
	data_exames_laboratorio_cap = request.POST.get('data_exames_laboratorio')
	data_exames_laboratorio_bd_cap = request.POST.get('data_exames_laboratorio_bd')

	if data_exames_laboratorio_cap == '' or data_exames_laboratorio_cap == None:
		if data_exames_laboratorio_bd_cap == '' or data_exames_laboratorio_bd_cap == None:
			data_exames_laboratorio = None
		else:
			data_exames_laboratorio = datetime.strptime(data_exames_laboratorio_bd_cap, '%d/%m/%Y').date()
	else:
		data_exames_laboratorio = data_exames_laboratorio_cap

	leucocito = request.POST.get('leucocito')
	linfocito = request.POST.get('linfocito')
	hb = request.POST.get('hb')
	ht = request.POST.get('ht')
	pcr = request.POST.get('pcr')
	lactato = request.POST.get('lactato')
	dhl = request.POST.get('dhl')
	ferritina = request.POST.get('ferritina')
	troponina = request.POST.get('troponina')
	cpk = request.POST.get('cpk')
	d_dimero = request.POST.get('d_dimero')
	ureia = request.POST.get('ureia')
	creatinina = request.POST.get('creatinina')
	ap = request.POST.get('ap')
	tap = request.POST.get('tap')
	inr = request.POST.get('inr')
	tgo = request.POST.get('tgo')
	tgp = request.POST.get('tgp')
	exame_imagem = request.POST.getlist('exame_imagem')
	#laudo_tc_torax = request.POST.get('laudo_tc_torax')
	#laudo_rx_torax = request.POST.get('laudo_rx_torax')
	is_sindrome_gripal = request.POST.get('sindrome_gripal')
	news_fast_pb = request.POST.get('new_fast')
	news_modificado = request.POST.get('new_modificado')
	uti = request.POST.get('uti')
	leito = request.POST.get('perfil')

	parecer_medico = request.POST.get('parecer')
	
	prioridade_cap = request.POST.get('prioridade')
	if prioridade_cap == '' or prioridade_cap == None:
		prioridade = None
	else:
		prioridade = int(float(prioridade_cap))

	last_status = request.POST.get('status_paciente')
	

	registro = RegistroCovid.objects.create(
		responsavel_pelo_preenchimento = responsavel_pelo_preenchimento,
		codigo_registro_total = codigo_registro_total,
		codigo_registro_mensal = codigo_registro_mensal,
		codigo_registro_completo = codigo_registro_completo,
		nome_solicitante = nome_solicitante,
		municipio_estabelecimento_solicitante = municipio_estabelecimento_solicitante,
		estabelecimento_solicitante = estabelecimento_solicitante,
		estabelecimento_solicitante_outro = estabelecimento_solicitante_outro,
		nome_paciente = nome_paciente,
		idade_paciente = idade_paciente,
		sexo_paciente = sexo_paciente,
		recurso_que_precisa = recurso_que_precisa,
		estado_origem = estado_origem,
		cidade_origem = cidade_origem,
		telefone_retorno = telefone_retorno,
		data_admissao = data_admissao,
		frequencia_respiratoria = frequencia_respiratoria,
		saturacao_paciente = saturacao_paciente,
		ar_o2 = ar_o2,
		frequencia_cardiaca = frequencia_cardiaca,
		pa_admissao = pa_admissao,
		pa = pa,
		conciencia = conciencia,
		temperatura = temperatura,
		image_descricao_clinica = image_descricao_clinica,
		image_laudo_tc = image_laudo_tc,
		image_laudo_rx = image_laudo_rx,
		observacao = observacao,
		descricao_clinica = descricao_clinica,
		sindrome_gripal = sindrome_gripal,
		tempo_quadro_sintomatico = tempo_quadro_sintomatico,
		exposicao_pessoa_infectada = exposicao_pessoa_infectada,
		parentesco = parentesco,
		comorbidades = comorbidades,
		outras_comorbidades = outras_comorbidades,
		outras_cardiopatias = outras_cardiopatias,
		idade_gestacional = idade_gestacional,
		comorbidades_obstetricas = comorbidades_obstetricas,
		gesta_para = gesta_para,
		medicamentos_uso_habitual = medicamentos_uso_habitual,
		medicamento_outros = medicamento_outros,
		spo2 = spo2,
		fr_irpm = fr_irpm,
		ventilacao_tipo = ventilacao_tipo,
		o2_suporte = o2_suporte,
		dose_cn = dose_cn,
		dose_venturi = dose_venturi,
		vmi = vmi,
		vt = vt,
		delta_pressure = delta_pressure,
		pplato = pplato,
		raw = raw,
		cest = cest,
		sao2 = sao2,
		pao2 = pao2,
		fio2 = fio2,
		paco2 = paco2,
		fc = fc,
		temperatura_axilar = temperatura_axilar,
		droga_vasoativa = droga_vasoativa,
		qtd_nora = qtd_nora,
		qtd_adrenalina = qtd_adrenalina,
		qtd_vasopressina = qtd_vasopressina,
		qtd_dobutamina = qtd_dobutamina,
		arritmia = arritmia,
		infeccao_bacteriana = infeccao_bacteriana,
		foco = foco,
		uso_antibioticoterapia = uso_antibioticoterapia,
		pesquisa_teste_sars_cov_2 = pesquisa_teste_sars_cov_2,
		igg = igg,
		igm = igm,
		data_igm_igg = data_igm_igg,
		rt_pcr_sars_cov_2 = rt_pcr_sars_cov_2,
		data_coleta = data_coleta,
		em_uso_corticosteroide = em_uso_corticosteroide,
		dose_corticosteroide = dose_corticosteroide,
		data_inicio_corticosteroide = data_inicio_corticosteroide,
		em_uso_hidroxicloroquina = em_uso_hidroxicloroquina,
		data_inicio_hidroxicloroquina = data_inicio_hidroxicloroquina,
		em_uso_oseltamivir = em_uso_oseltamivir,
		data_inicio_oseltamivir = data_inicio_oseltamivir,
		em_uso_heparina = em_uso_heparina,
		data_inicio_heparina = data_inicio_heparina,
		tipo_heparina = tipo_heparina,
		dose_heparina = dose_heparina,
		pps = pps,
		escala_pontos_glasgow = escala_pontos_glasgow,
		bloqueador_neuromuscular = bloqueador_neuromuscular,
		midazolam_dose = midazolam_dose,
		fentanil_dose = fentanil_dose,
		propofol_dose = propofol_dose,
		sedacao_continua = sedacao_continua,
		rocuronio_dose = rocuronio_dose,
		pacuronio_dose = pacuronio_dose,
		cisatracurio_dose = cisatracurio_dose,
		rass = rass,
		data_exames_laboratorio = data_exames_laboratorio,
		leucocito = leucocito,
		linfocito = linfocito,
		hb = hb,
		ht = ht,
		pcr = pcr,
		lactato = lactato,
		dhl = dhl,
		ferritina = ferritina,
		troponina = troponina,
		cpk = cpk,
		d_dimero = d_dimero,
		ureia = ureia,
		creatinina = creatinina,
		ap = ap,
		tap = tap,
		inr = inr,
		tgo = tgo,
		tgp = tgp,
		exame_imagem = exame_imagem,
		#laudo_tc_torax = laudo_tc_torax,
		#laudo_rx_torax = laudo_rx_torax,
		is_sindrome_gripal = is_sindrome_gripal,
		news_fast_pb = news_fast_pb,
		news_modificado = news_modificado,
		uti = uti,
		leito = leito,
		parecer_medico = parecer_medico,
		prioridade = prioridade,
		last_status = last_status
		#senha = senha,
		#codigo_sescovid = codigo_sescovid,
		#justificativa = justificativa,
		#pareceristas = pareceristas,
		#data_regulacao = data_regulacao,
		#data_obito = data_obito

		)

	return redirect('/covid_list/')

@login_required
def registro_enfermeiro_medico(request):
    return render(request, 'registro_enfermeiro_medico.html')


@login_required
def regulacao_search_estabelecimento(request, id):
	registro = RegistroCovid.objects.get(id=id)
	municipios = Cnes.objects.order_by('MUNICIPIO').distinct('MUNICIPIO')

	return render(request, 'regulacao_search_estabelecimento.html', {'registro':
		registro, 'municipios':municipios})


@login_required
def regulacao_search_estabelecimento_set(request, id):
	municipio_estabelecimento_referencia = request.session['municipio_estabelecimento_referencia']

	return redirect('paciente_atribuir_senha', id=id)


@login_required
def regulacao(request, id):
	data_regulacao = datetime.now()
	data_regulacao_template = data_regulacao.strftime('%d/%m/%Y %H:%M')

	registro = RegistroCovid.objects.get(id=id)

	pareceristas_bd = registro.pareceristas
	pareceristas = []
	if pareceristas_bd:
		pareceristas = list(set(pareceristas_bd))

	#registro_last = RegistroCovid.objects.all().last()
	#senha_last = registro_last.senha
	
	#Regulações do paciente
	regulacoes_registro = Regulacao.objects.filter(registro_covid_id=registro.id)

	#ultima regulacao
	regulacao_last = Regulacao.objects.all().last()
	#Ultima regulacao do paciente:
	regulacao_paciente_last = regulacoes_registro.last()
	
	if regulacao_last:
		senha_last = regulacao_last.senha
		senha_new = senha_last + 1
	else:
		senha_new = 500000


	senha = senha_new

	codigo_sescovid = "SESCOVID" + str(senha_new)
	

	pa = registro.pa

	p = pa.split("x")
	pa_1 = p[0]
	pa_2 = p[1]


	pa_admissao = registro.pa_admissao
	p = pa_admissao.split("x")
	pad_1 = p[0]
	pad_2 = p[1]

	status_registro = Status.objects.filter(registro_covid=registro.id)

	status_list_descricao_ = []

	for s in status_registro:
		status_list_descricao_.append(s.descricao)



	status_list_descricao = []
	if len(status_list_descricao_) != 0:
		status_list_descricao = status_list_descricao_.pop()
	else:
		status_list_descricao
	status_aguard_conf_vaga_registro = Status.objects.filter(descricao='Aguardando confirmação de Vaga').last()
	#status_obito_registro = Status.objects.filter(descricao='Obito').last()
	status_aguard_lista_espera_registro = Status.objects.filter(descricao='Aguardando em Lista de Espera').last()
	status_regulado_registro = Status.objects.filter(descricao='Regulado').last()
	
	#status_nao_regulado_registro = []
	status_nao_regulado_registro = Status.objects.filter(descricao='Não Regulado').last()
	#status_nao_regulado = status_nao_regulado_registro.last()


	municipio_estabelecimento_referencia = request.POST.get('municipio_estabelecimento_referencia')
	#request.session['nome_solicitante'] = nome_solicitante
	#request.session['idade_paciente'] = idade_paciente
	#request.session['nome_paciente'] = nome_paciente
	#request.session['sexo_paciente'] = sexo_paciente
	municipio_estabelecimento_solicitante = request.POST.get('municipio_estabelecimento_solicitante')
	estabelecimento_solicitante = request.POST.get('estabelecimento_solicitante')
	estabelecimento_solicitante_outro = request.POST.get('estabelecimento_solicitante_outro')
	#request.session['recurso_que_precisa'] = recurso_que_precisa
	#request.session['cidade_origem'] = cidade_origem
	#request.session['telefone_retorno'] = telefone_retorno
	#request.session['frequencia_respiratoria'] = frequencia_respiratoria
	#request.session['saturacao_paciente'] = saturacao_paciente
	#request.session['frequencia_cardiaca'] = frequencia_cardiaca
	#request.session['conciencia'] = conciencia
	#request.session['temperatura'] = temperatura
	#request.session['observacao'] = observacao
	#request.session['pa_1'] = pa_1
	#request.session['pa_2'] = pa_2
	cidade = request.POST.get('municipio_referencia')


	estabelecimentos = Cnes.objects.filter(MUNICIPIO = municipio_estabelecimento_referencia).order_by('NO_FANTASIA')

	encoded_ = None
	if registro.image_descricao_clinica:
		encoded_ = base64.b64encode(registro.image_descricao_clinica).decode('ascii')

	encoded = encoded_

	if not registro.image_laudo_tc and not registro.image_laudo_rx:
		return render(request, 'regulacao.html', {'registro' : registro, 
			'codigo_sescovid':codigo_sescovid, 'senha':senha, 'pa_1':pa_1, 
			'pa_2':pa_2, 'pad_1':pad_1, 'pad_2':pad_2,
			'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro,
			'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'pareceristas':pareceristas,
			'status_regulado_registro':status_regulado_registro, 
			'status_nao_regulado_registro':status_nao_regulado_registro, 
			'municipio_estabelecimento_solicitante':
			municipio_estabelecimento_solicitante, 
			'estabelecimento_solicitante':estabelecimento_solicitante, 
			'estabelecimento_solicitante_outro':
			estabelecimento_solicitante_outro, 'cidade':cidade,
			'estabelecimentos':estabelecimentos,
			'municipio_estabelecimento_referencia':
			municipio_estabelecimento_referencia, 'encoded':encoded})
	elif not registro.image_laudo_tc and not registro.image_laudo_rx and not registro.image_descricao_clinica:
		return render(request, 'regulacao.html', {'registro' : registro, 
			'codigo_sescovid':codigo_sescovid, 'senha':senha, 'pa_1':pa_1, 
			'pa_2':pa_2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro,
			'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 
			'status_regulado_registro':status_regulado_registro, 
			'status_nao_regulado_registro':status_nao_regulado_registro, 
			'municipio_estabelecimento_solicitante':
			municipio_estabelecimento_solicitante, 
			'estabelecimento_solicitante':estabelecimento_solicitante, 
			'estabelecimento_solicitante_outro':
			estabelecimento_solicitante_outro, 'cidade':cidade,
			'estabelecimentos':estabelecimentos,
			'municipio_estabelecimento_referencia':
			municipio_estabelecimento_referencia, 'pad_1':pad_1, 'pad_2':pad_2, 
			'pareceristas':pareceristas,})

	elif not registro.image_laudo_rx:
		img_tc = base64.b64encode(registro.image_laudo_tc).decode('ascii')

		return render(request, 'regulacao.html', {'registro' : registro, 
		'codigo_sescovid':codigo_sescovid, 'senha':senha, 'pa_1':pa_1, 
		'pa_2':pa_2, 'data_regulacao_template': data_regulacao_template,
		'status_list_descricao':status_list_descricao, 'status_aguard_conf_vaga_registro':
		status_aguard_conf_vaga_registro, 'status_aguard_lista_espera_registro':
		status_aguard_lista_espera_registro, 'status_regulado_registro':
		status_regulado_registro, 'status_nao_regulado_registro':
		status_nao_regulado_registro, 'municipio_estabelecimento_solicitante':
		municipio_estabelecimento_solicitante, 'estabelecimento_solicitante':
		estabelecimento_solicitante, 'estabelecimento_solicitante_outro':
		estabelecimento_solicitante_outro, 'cidade':cidade, 'estabelecimentos':
		estabelecimentos, 'municipio_estabelecimento_referencia':
		municipio_estabelecimento_referencia, 'encoded':encoded, 'img_tc':
		img_tc, 'pad_1':pad_1, 'pad_2':pad_2, 'pareceristas':pareceristas,})
	elif not registro.image_laudo_tc:
		img_rx = base64.b64encode(registro.image_laudo_rx).decode('ascii')

		return render(request, 'regulacao.html', {'registro' : registro, 
		'codigo_sescovid':codigo_sescovid, 'senha':senha, 'pa_1':pa_1, 
		'pa_2':pa_2, 'data_regulacao_template': data_regulacao_template,
		'status_list_descricao':status_list_descricao, 'status_aguard_conf_vaga_registro':
		status_aguard_conf_vaga_registro, 'status_aguard_lista_espera_registro':
		status_aguard_lista_espera_registro, 'status_regulado_registro':
		status_regulado_registro, 'status_nao_regulado_registro':
		status_nao_regulado_registro, 'municipio_estabelecimento_solicitante':
		municipio_estabelecimento_solicitante, 'estabelecimento_solicitante':
		estabelecimento_solicitante, 'estabelecimento_solicitante_outro':
		estabelecimento_solicitante_outro, 'cidade':cidade, 'estabelecimentos':
		estabelecimentos, 'municipio_estabelecimento_referencia':
		municipio_estabelecimento_referencia, 'encoded':encoded, 'img_rx':img_rx,
		 'pad_1':pad_1, 'pad_2':pad_2, 'pareceristas':pareceristas,})
	else:
		img_tc = base64.b64encode(registro.image_laudo_tc).decode('ascii')
		img_rx = base64.b64encode(registro.image_laudo_rx).decode('ascii')

		return render(request, 'regulacao.html', {'registro' : registro,
		'codigo_sescovid':codigo_sescovid, 'senha':senha, 'pa_1':pa_1, 'pa_2':pa_2, 
		'data_regulacao_template': data_regulacao_template,
		'status_list_descricao':status_list_descricao, 'status_aguard_conf_vaga_registro':
		status_aguard_conf_vaga_registro, 'status_aguard_lista_espera_registro':
		status_aguard_lista_espera_registro, 'status_regulado_registro':
		status_regulado_registro, 'status_nao_regulado_registro':
		status_nao_regulado_registro, 'municipio_estabelecimento_solicitante':
		municipio_estabelecimento_solicitante, 'estabelecimento_solicitante':
		estabelecimento_solicitante, 'estabelecimento_solicitante_outro':
		estabelecimento_solicitante_outro, 'cidade':cidade, 'estabelecimentos':
		estabelecimentos, 'municipio_estabelecimento_referencia':
		municipio_estabelecimento_referencia, 'encoded':encoded, 'img_tc':
		img_tc, 'img_rx':img_rx, 'pad_1':pad_1, 'pad_2':pad_2, 
		'pareceristas':pareceristas,})





	

@login_required
def regulacao_set(request, id):
	registro = RegistroCovid.objects.get(id=id)

	responsavel_pelo_preenchimento = request.user
	
	
	nome_solicitante = request.POST.get('nome_solicitante')
	municipio_estabelecimento_solicitante = registro.municipio_estabelecimento_solicitante
	estabelecimento_solicitante = registro.estabelecimento_solicitante
	estabelecimento_solicitante_outro = registro.estabelecimento_solicitante_outro
	
	if estabelecimento_solicitante == '' or estabelecimento_solicitante == None:
		estabelecimento_solicitante = request.POST.get('estabelecimento_solicitante')
	
	if estabelecimento_solicitante_outro == '' or estabelecimento_solicitante_outro == None:
		estabelecimento_solicitante_outro = request.POST.get('estabelecimento_solicitante_outro')

	municipio_estabelecimento_referencia = request.POST.get('municipio_estabelecimento_referencia')
	estabelecimento_referencia = request.POST.get('estabelecimento_referencia')
	estabelecimento_referencia_outro = request.POST.get('estabelecimento_referencia_outro')
	nome_paciente = request.POST.get('nome_paciente')
	
	idade_paciente_cap = request.POST.get('idade_paciente')
	if idade_paciente_cap == '' or idade_paciente_cap == None:
		idade_paciente = None
	else:
		idade_paciente = int(float(idade_paciente_cap))

	sexo_paciente = request.POST.get('sexo_paciente')
	recurso_que_precisa = registro.recurso_que_precisa
	estado_origem = registro.estado_origem
	cidade_origem = registro.cidade_origem
	telefone_retorno = registro.telefone_retorno
	
	data_admissao_cap = request.POST.get('data_admissao')

	if data_admissao_cap != '' or data_admissao_cap != None:
		data_admissao = datetime.strptime(data_admissao_cap, '%d/%m/%Y').date()
		#data_admissao = data_admissao_cap


	frequencia_respiratoria_cap = request.POST.get('frequencia_respiratoria')
	if frequencia_respiratoria_cap == '' or frequencia_respiratoria_cap == None:
		frequencia_respiratoria = registro.frequencia_respiratoria
	else:
		frequencia_respiratoria = int(frequencia_respiratoria_cap)
		
	
	saturacao_paciente_cap = request.POST.get('saturacao_paciente')
	if saturacao_paciente_cap != '':
		saturacao_paciente = int(saturacao_paciente_cap)
	else:
		saturacao_paciente = registro.saturacao_paciente


	frequencia_cardiaca_cap = request.POST.get('f_cardiaca_paciente')
	if frequencia_cardiaca_cap != '':
		frequencia_cardiaca = int(frequencia_cardiaca_cap)
	else:
		frequencia_cardiaca = registro.frequencia_cardiaca

	
	ar_o2_cap = request.POST.get('ar_o2')
	if ar_o2_cap != '' or ar_o2_cap != None:
		ar_o2 = ar_o2_cap
	else:
		ar_o2 = registro.ar_o2
	
	
	pa_part1 = request.POST.get('pa_part1')
	pa_part2 = request.POST.get('pa_part2')
	pa = pa_part1 + "x" + pa_part2

	pad_part1 = request.POST.get('pa_admissao1')
	pad_part2 = request.POST.get('pa_admissao2')
	pa_admissao = pad_part1 + "x" + pad_part2
	
	conciencia_cap = request.POST.get('consciencia_paciente')
	if conciencia_cap != '' or conciencia_cap != None:
		conciencia = conciencia_cap
	else:
		conciencia = registro.conciencia

	temperatura_cap = request.POST.get('temperatura_paciente')
	if temperatura_cap == '' or temperatura_cap == None:
		temperatura = registro.temperatura
	else:
		temperatura = float(temperatura_cap)
		

	descricao_clinica = request.POST.get('descricao_clinica')

	#image_descricao_clinica = request.FILES['file_descricao_clinica'].file.read()

	#image_laudo_tc = request.FILES['file_tc_torax_regulacao'].file.read()
	#image_laudo_rx = request.FILES['file_rx_torax_regulacao'].file.read()
	
	sindrome_gripal = request.POST.getlist('s_gripal')
	
	tempo_quadro_sintomatico_cap = request.POST.get('tempo_inicio_sintomas')
	if tempo_quadro_sintomatico_cap == '' or tempo_quadro_sintomatico_cap == None:
		tempo_quadro_sintomatico = None
	else:
		tempo_quadro_sintomatico = int(float(tempo_quadro_sintomatico_cap))

	exposicao_pessoa_infectada = request.POST.get('exposicao_epidemiologica_sim')
	parentesco = request.POST.get('grau_parentesco')
	comorbidades = request.POST.getlist('comorbidades')
	outras_comorbidades = request.POST.get('descricao_outras_comorbidades')
	outras_cardiopatias = request.POST.get('desc_outras_cardiopatias')
	idade_gestacional = request.POST.get('idade_gestacao')
	comorbidades_obstetricas = request.POST.get('comorbidade_obstetrica')
	gesta_para = request.POST.get('gestacao_para')
	medicamentos_uso_habitual = request.POST.getlist('medicamento_uso_habitual')
	medicamento_outros = request.POST.get('descricao_outros_medicamento')
	
	spo2_cap = request.POST.get('spo2')
	if spo2_cap == '' or spo2_cap == None:
		spo2 = None
	else:
		spo2 = int(float(spo2_cap))
		

	fr_irpm_cap = request.POST.get('fr_irpm')
	if fr_irpm_cap == '' or fr_irpm_cap == None:
		fr_irpm = None
	else:
		fr_irpm = int(float(fr_irpm_cap))
		

	ventilacao_tipo = request.POST.get('ventilacao')
	o2_suporte = request.POST.getlist('o2_suporte')
	
	dose_cn_cap = request.POST.get('dose_cn')
	if dose_cn_cap == '' or dose_cn_cap == None:
		dose_cn = None
	else:
		dose_cn = float(dose_cn_cap)

	dose_venturi_cap = request.POST.get('dose_venturi')
	if dose_venturi_cap == '' or dose_venturi_cap == None:
		dose_venturi = None
	else:
		dose_venturi = float(dose_venturi_cap)
	

	vmi = request.POST.getlist('vmi')
	
	vt_cap = request.POST.get('vt')
	if vt_cap == '' or vt_cap == None:
		vt = None
	else:
		vt = float(vt_cap)
		

	delta_pressure_cap = request.POST.get('delta_pressure')
	if delta_pressure_cap == '' or delta_pressure_cap == None:
		delta_pressure = None
	else:
		delta_pressure = float(delta_pressure_cap)
		

	pplato_cap = request.POST.get('pplato')
	if pplato_cap == '' or pplato_cap == None:
		pplato = None
	else:
		pplato = float(pplato_cap)
		

	raw_cap = request.POST.get('raw')
	if raw_cap == '' or raw_cap == None:
		raw = None
	else:
		raw = float(raw_cap)
		

	cest_cap = request.POST.get('cest')
	if cest_cap == '' or cest_cap == None :
		cest = None
	else:
		cest = float(cest_cap)
		

	sao2_cap = request.POST.get('sao2')
	if sao2_cap == '' or sao2_cap == None:
		sao2 = None
	else:
		sao2 = float(sao2_cap)

	pao2_cap = request.POST.get('pao2')
	if pao2_cap == '' or pao2_cap == None:
		pao2 = None
	else:
		pao2 = float(pao2_cap)

	fio2_cap = request.POST.get('fio2')
	if fio2_cap == '' or fio2_cap == None:
		fio2 = None
	else:
		fio2 = float(fio2_cap)

	paco2_cap = request.POST.get('paco2')
	if paco2_cap == '' or paco2_cap == None:
		paco2 = None
	else:
		paco2 = float(paco2_cap)
	
	fc_cap = request.POST.get('fc')
	if fc_cap == '' or fc_cap == None:
		fc = None
	else:
		fc = int(float(fc_cap))
		

	temperatura_axilar_cap = request.POST.get('temp_auxiliar')
	if temperatura_axilar_cap == '' or temperatura_axilar_cap == None :
		temperatura_axilar = None
	else:
		temperatura_axilar = float(temperatura_axilar_cap)
		

	droga_vasoativa = request.POST.getlist('droga_vasoativa')
	
	qtd_nora_cap = request.POST.get('qtd_nora')
	if qtd_nora_cap == '' or qtd_nora_cap == None:
		qtd_nora = None
	else:
		qtd_nora = float(qtd_nora_cap)

	qtd_adrenalina_cap = request.POST.get('qtd_adrenalina')
	if qtd_adrenalina_cap == '' or qtd_adrenalina_cap == None:
		qtd_adrenalina = None
	else:
		qtd_adrenalina = float(qtd_adrenalina_cap)

	qtd_vasopressina_cap = request.POST.get('qtd_vasopressina')
	if qtd_vasopressina_cap == '' or qtd_vasopressina_cap == None:
		qtd_vasopressina = None
	else:
		qtd_vasopressina = float(qtd_vasopressina_cap)

	qtd_dobutamina_cap = request.POST.get('qtd_dobutamina')
	if qtd_dobutamina_cap == '' or qtd_dobutamina_cap == None:
		qtd_dobutamina = None
	else:
		qtd_dobutamina = float(qtd_dobutamina_cap)

	arritmia = request.POST.getlist('arritmia')
	infeccao_bacteriana = request.POST.get('infeccao_bacteriana')
	foco = request.POST.getlist('foco')
	uso_antibioticoterapia = request.POST.get('uso_antibioticoterapia')
	pesquisa_teste_sars_cov_2 = request.POST.get('pesquisa_sars_cov2')
	igg_cap = request.POST.get('igg')
	igg_bd = request.POST.get('igg_bd')
	if igg_cap == '' or igg_cap == None:
		if igg_bd == '' or igg_bd == None:
			igg = None
		else:
			igg = igg_bd
	else:
		igg = igg_cap

	igm_do_bd = registro.igm
	igm_cap = request.POST.get('igm')

	igm_bd_cap = request.POST.get('igm_bd')
	if igm_cap == '' or igm_cap == None:
		if igm_bd_cap == '' or igm_bd_cap == None:
			igm = None
		else:
			igm = igm_bd_cap
	else:
		igm = igm_cap

	
	data_igm_igg_cap = request.POST.get('data_igm_igg')
	data_igm_igg_bd_cap = request.POST.get('data_igm_igg_bd')

	if data_igm_igg_cap == '' or data_igm_igg_cap == None:
		if data_igm_igg_bd_cap == '' or data_igm_igg_bd_cap == None:
			data_igm_igg = None
		else:
			data_igm_igg = datetime.strptime(data_igm_igg_bd_cap, '%d/%m/%Y').date()
	else:
		data_igm_igg = data_igm_igg_cap


	rt_pcr_sars_cov_2 = request.POST.get('pcr_sars_cov2')
	
	data_da_coleta_bd_cap = request.POST.get('data_da_coleta_bd')
	data_coleta_cap = request.POST.get('data_da_coleta')
	if data_coleta_cap == '' or data_coleta_cap == None:
		if data_da_coleta_bd_cap == '' or data_da_coleta_bd_cap == None:
			data_coleta = None
		else:
			data_coleta = datetime.strptime(data_da_coleta_bd_cap, '%d/%m/%Y').date()
	else:
		data_coleta = data_coleta_cap
	
	em_uso_corticosteroide = request.POST.get('em_uso_corticosteroide')
	dose_corticosteroide = request.POST.get('dose_corticosteroide')
	
	data_inicio_corticosteroide_bd_cap = request.POST.get('data_inicio_corticosteroide_bd')
	data_inicio_corticosteroide_cap = request.POST.get('data_inicio_corticosteroide')
	if data_inicio_corticosteroide_cap == '' or data_inicio_corticosteroide_cap == None:
		if data_inicio_corticosteroide_bd_cap == '' or data_inicio_corticosteroide_bd_cap == None:
			data_inicio_corticosteroide = None
		else:
			data_inicio_corticosteroide = datetime.strptime(data_inicio_corticosteroide_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_corticosteroide = data_inicio_corticosteroide_cap
	

	em_uso_hidroxicloroquina = request.POST.get('em_uso_hidrocloroquina')
	
	data_inicio_hidroxicloroquina_cap = request.POST.get('data_inicio_hidroxicloroquina')
	data_inicio_hidroxicloroquina_bd_cap = request.POST.get('data_inicio_hidroxicloroquina_bd')


	if data_inicio_hidroxicloroquina_cap == '' or data_inicio_hidroxicloroquina_cap == None:
		if data_inicio_hidroxicloroquina_bd_cap == '' or data_inicio_hidroxicloroquina_bd_cap == None:
			data_inicio_hidroxicloroquina = None	
		else:
			data_inicio_hidroxicloroquina = datetime.strptime(data_inicio_hidroxicloroquina_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_hidroxicloroquina = data_inicio_hidroxicloroquina_cap


	

	em_uso_oseltamivir = request.POST.get('em_uso_oseltamivir')
	
	data_inicio_oseltamivir_cap = request.POST.get('data_inicio_oseltamivir')
	data_inicio_oseltamivir_bd_cap = request.POST.get('data_inicio_oseltamivir_bd')


	if data_inicio_oseltamivir_cap == '' or data_inicio_oseltamivir_cap == None:
		if data_inicio_oseltamivir_bd_cap == '' or data_inicio_oseltamivir_bd_cap == None:
			data_inicio_oseltamivir = None
		else:
			data_inicio_oseltamivir = datetime.strptime(data_inicio_oseltamivir_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_oseltamivir = data_inicio_oseltamivir_cap



	em_uso_heparina = request.POST.get('em_uso_heparina')

	data_inicio_heparina_bd_cap = request.POST.get('data_inicio_heparina_bd')
	data_inicio_heparina_cap = request.POST.get('data_inicio_heparina')

	if data_inicio_heparina_cap == '' or data_inicio_heparina_cap == None:
		if data_inicio_heparina_bd_cap == '' or data_inicio_heparina_bd_cap == None:
			data_inicio_heparina = None
		else:
			data_inicio_heparina = datetime.strptime(data_inicio_heparina_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_heparina = data_inicio_heparina_cap
		

	tipo_heparina = request.POST.get('tipo_heparina')
	dose_heparina = request.POST.get('dose_heparina')
	
	pps_cap = request.POST.get('pps_paciente')
	if pps_cap == '' or pps_cap == None:
		pps = None
	else:
		pps = float(pps_cap)
		

	escala_pontos_glasgow_cap = request.POST.get('escala_glasgow')
	if escala_pontos_glasgow_cap == '' or escala_pontos_glasgow_cap == None:
		escala_pontos_glasgow = None
	else:
		escala_pontos_glasgow = int(float(escala_pontos_glasgow_cap))
		

	bloqueador_neuromuscular = request.POST.getlist('desc_bloqueador_neuromuscular')
	midazolam_dose = request.POST.get('midazolam_dose')
	fentanil_dose = request.POST.get('fentanil_dose')
	propofol_dose = request.POST.get('propofol_dose')
	sedacao_continua = request.POST.getlist('desc_sedacao_continua')
	rocuronio_dose = request.POST.get('rocuronio_dose')
	pacuronio_dose = request.POST.get('pacuronio_dose')
	cisatracurio_dose = request.POST.get('cisatracurio_dose')
	rass = request.POST.get('rass')
	
	data_exames_laboratorio_cap = request.POST.get('data_exames_laboratorio')
	data_exames_laboratorio_bd_cap = request.POST.get('data_exames_laboratorio_bd')

	if data_exames_laboratorio_cap == '' or data_exames_laboratorio_cap == None:
		if data_exames_laboratorio_bd_cap == '' or data_exames_laboratorio_bd_cap == None:
			data_exames_laboratorio = None
		else:
			data_exames_laboratorio = datetime.strptime(data_exames_laboratorio_bd_cap, '%d/%m/%Y').date()
	else:
		data_exames_laboratorio = data_exames_laboratorio_cap

	leucocito = request.POST.get('leucocito')
	linfocito = request.POST.get('linfocito')
	hb = request.POST.get('hb')
	ht = request.POST.get('ht')
	pcr = request.POST.get('pcr')
	lactato = request.POST.get('lactato')
	dhl = request.POST.get('dhl')
	ferritina = request.POST.get('ferritina')
	troponina = request.POST.get('troponina')
	cpk = request.POST.get('cpk')
	d_dimero = request.POST.get('d_dimero')
	ureia = request.POST.get('ureia')
	creatinina = request.POST.get('creatinina')
	ap = request.POST.get('ap')
	tap = request.POST.get('tap')
	inr = request.POST.get('inr')
	tgo = request.POST.get('tgo')
	tgp = request.POST.get('tgp')
	exame_imagem = request.POST.getlist('exame_imagem')
	laudo_tc_torax = request.POST.get('laudo_tc_torax')
	laudo_rx_torax = request.POST.get('laudo_rx_torax')
	is_sindrome_gripal = request.POST.get('sindrome_gripal')
	news_fast_pb = request.POST.get('new_fast')
	news_modificado = request.POST.get('new_modificado')
	uti = request.POST.getlist('uti')
	leito = request.POST.get('perfil')

	parecer_medico = request.POST.get('parecer')
	
	prioridade_cap = request.POST.get('prioridade')
	if prioridade_cap == '' or prioridade_cap == None:
		prioridade = None
	else:
		prioridade = int(float(prioridade_cap))


	descricao = request.POST.get('status_paciente')
	registro_covid = registro

	status = Status.objects.create(
		descricao=descricao,
		registro_covid=registro_covid
		)

	last_status = descricao

	if estabelecimento_referencia == '' or estabelecimento_referencia == None:
		estabelecimento_referencia_covid = estabelecimento_referencia_outro
	else:
		estabelecimento_referencia_covid = estabelecimento_referencia


	codigo_sescovid = ""
	senha = 0
	if last_status == 'Regulado':
		#Regulações do paciente
		#regulacoes_registro = Regulacao.objects.filter(registro_covid_id=registro.id)
		#Ultima regulacao do paciente:
		codigo_sescovid = request.POST.get('num_protocolo')
		regulacao_last = Regulacao.objects.all().last()
		if regulacao_last:
			senha_last = regulacao_last.senha
			senha_new = senha_last + 1
		else:
			senha_new = 500000
			

		senha = int(senha_new)

		regulacao = Regulacao.objects.create(
			estabelecimento_referencia_covid = estabelecimento_referencia_covid,
			nome_paciente = nome_paciente,
			senha = senha,
			registro_covid = registro_covid
			)


	justificativa = request.POST.get('justificativa')
	observacao = request.POST.get('observacoes_medicas')
	
	pareceristas = registro.pareceristas
	
	pareceristas_cap = request.POST.getlist('pareceristas')
	
	if pareceristas == '' or pareceristas == None:
		pareceristas = pareceristas_cap
	else:
		pareceristas = pareceristas + pareceristas_cap

	
	data_regulacao = datetime.now()
	#data_regulacao = data_regulacao_.strftime('%d/%m/%Y %H:%M')


	data_obito_bd_cap = request.POST.get('data_obito_bd')
	data_obito_cap = request.POST.get('data_obito')

	if data_obito_cap == '' or data_obito_cap == None:
		if data_obito_bd_cap == '' or data_obito_bd_cap == None:
			data_obito = None
		else:
			data_obito = datetime.strptime(data_obito_bd_cap, '%d/%m/%Y').date()
	else:
		data_obito = data_obito_cap

	
	registro.responsavel_pelo_preenchimento = responsavel_pelo_preenchimento
	#registro.codigo_registro = codigo_registro
	registro.nome_solicitante = nome_solicitante
	registro.municipio_estabelecimento_solicitante = municipio_estabelecimento_solicitante
	registro.estabelecimento_solicitante = estabelecimento_solicitante
	registro.estabelecimento_solicitante_outro = estabelecimento_solicitante_outro
	registro.municipio_estabelecimento_referencia = municipio_estabelecimento_referencia
	registro.estabelecimento_referencia = estabelecimento_referencia
	registro.estabelecimento_referencia_outro = estabelecimento_referencia_outro
	registro.nome_paciente = nome_paciente
	registro.idade_paciente = idade_paciente
	registro.sexo_paciente = sexo_paciente
	registro.recurso_que_precisa = recurso_que_precisa
	registro.estado_origem = estado_origem
	registro.cidade_origem = cidade_origem
	registro.telefone_retorno = telefone_retorno
	registro.data_admissao = data_admissao
	registro.frequencia_respiratoria = frequencia_respiratoria
	registro.saturacao_paciente = saturacao_paciente
	registro.ar_o2 = ar_o2
	registro.frequencia_cardiaca = frequencia_cardiaca
	registro.pa = pa
	registro.pa_admissao = pa_admissao
	registro.conciencia = conciencia
	registro.temperatura = temperatura
	registro.descricao_clinica = descricao_clinica
	#registro.image_descricao_clinica = image_descricao_clinica
	registro.sindrome_gripal = sindrome_gripal
	registro.tempo_quadro_sintomatico = tempo_quadro_sintomatico
	registro.exposicao_pessoa_infectada = exposicao_pessoa_infectada
	registro.parentesco = parentesco
	registro.comorbidades = comorbidades
	registro.outras_comorbidades = outras_comorbidades
	registro.outras_cardiopatias = outras_cardiopatias
	registro.idade_gestacional = idade_gestacional
	registro.comorbidades_obstetricas = comorbidades_obstetricas
	registro.gesta_para = gesta_para
	registro.medicamentos_uso_habitual = medicamentos_uso_habitual
	registro.medicamento_outros = medicamento_outros
	registro.spo2 = spo2
	registro.fr_irpm = fr_irpm
	registro.ventilacao_tipo = ventilacao_tipo
	registro.o2_suporte = o2_suporte
	registro.dose_cn = dose_cn
	registro.dose_venturi = dose_venturi
	registro.vmi = vmi
	registro.vt = vt
	registro.delta_pressure = delta_pressure
	registro.pplato = pplato
	registro.raw = raw
	registro.cest = cest
	registro.sao2 = sao2
	registro.pao2 = pao2
	registro.fio2 = fio2
	registro.paco2 = paco2
	registro.fc = fc
	registro.temperatura_axilar = temperatura_axilar
	registro.droga_vasoativa = droga_vasoativa
	registro.qtd_nora = qtd_nora
	registro.qtd_adrenalina = qtd_adrenalina
	registro.qtd_vasopressina = qtd_vasopressina
	registro.qtd_dobutamina = qtd_dobutamina
	registro.arritmia = arritmia
	registro.infeccao_bacteriana = infeccao_bacteriana
	registro.foco = foco
	registro.uso_antibioticoterapia = uso_antibioticoterapia
	registro.pesquisa_teste_sars_cov_2 = pesquisa_teste_sars_cov_2
	registro.igg = igg
	registro.igm = igm
	registro.data_igm_igg = data_igm_igg
	registro.rt_pcr_sars_cov_2 = rt_pcr_sars_cov_2
	registro.data_coleta = data_coleta
	registro.em_uso_corticosteroide = em_uso_corticosteroide
	registro.dose_corticosteroide = dose_corticosteroide
	registro.data_inicio_corticosteroide = data_inicio_corticosteroide
	registro.em_uso_hidroxicloroquina = em_uso_hidroxicloroquina
	registro.data_inicio_hidroxicloroquina = data_inicio_hidroxicloroquina
	registro.em_uso_oseltamivir = em_uso_oseltamivir
	registro.data_inicio_oseltamivir = data_inicio_oseltamivir
	registro.em_uso_heparina = em_uso_heparina
	registro.data_inicio_heparina = data_inicio_heparina
	registro.tipo_heparina = tipo_heparina
	registro.dose_heparina = dose_heparina
	registro.pps = pps
	registro.escala_pontos_glasgow = escala_pontos_glasgow
	registro.bloqueador_neuromuscular = bloqueador_neuromuscular
	registro.midazolam_dose = midazolam_dose
	registro.fentanil_dose = fentanil_dose
	registro.propofol_dose = propofol_dose
	registro.sedacao_continua = sedacao_continua
	registro.rocuronio_dose = rocuronio_dose
	registro.pacuronio_dose = pacuronio_dose
	registro.cisatracurio_dose = cisatracurio_dose
	registro.rass = rass
	registro.data_exames_laboratorio = data_exames_laboratorio
	registro.leucocito = leucocito
	registro.linfocito = linfocito
	registro.hb = hb
	registro.ht = ht
	registro.pcr = pcr
	registro.lactato = lactato
	registro.dhl = dhl
	registro.ferritina = ferritina
	registro.troponina = troponina
	registro.cpk = cpk
	registro.d_dimero = d_dimero
	registro.ureia = ureia
	registro.creatinina = creatinina
	registro.ap = ap
	registro.tap = tap
	registro.inr = inr
	registro.tgo = tgo
	registro.tgp = tgp
	registro.exame_imagem = exame_imagem
	#registro.image_laudo_tc = image_laudo_tc
	#registro.image_laudo_rx = image_laudo_rx
	registro.laudo_tc_torax = laudo_tc_torax
	registro.laudo_rx_torax = laudo_rx_torax
	registro.is_sindrome_gripal = is_sindrome_gripal
	registro.news_fast_pb = news_fast_pb
	registro.news_modificado = news_modificado
	registro.uti = uti
	registro.leito = leito
	registro.parecer_medico = parecer_medico
	registro.prioridade = prioridade
	#registro.regulacao_paciente = regulacao_paciente
	#registro.status_regulacao = status_regulacao
	#registro.senha = senha
	registro.codigo_sescovid = codigo_sescovid
	registro.justificativa = justificativa
	registro.observacao = observacao
	registro.pareceristas = pareceristas
	registro.data_regulacao = data_regulacao
	registro.last_status = last_status
	registro.data_obito = data_obito

	registro.save()

	return redirect('regulacao_detail', id=id)


@login_required
def regulacao_detail(request, id):
	registro = RegistroCovid.objects.get(id=id)

	pareceristas_bd = registro.pareceristas
	pareceristas = []
	if pareceristas_bd:
		pareceristas = list(set(pareceristas_bd))

	regulacoes_registro = Regulacao.objects.filter(registro_covid_id=registro.id)
	regulacao_registro_last = regulacoes_registro.last()
	

	status_registro = Status.objects.filter(registro_covid=registro.id)
	status_registro_last = Status.objects.filter(registro_covid=registro.id).last()

	status_aguard_conf_vaga_registro = status_registro.filter(descricao='Aguardando confirmação de Vaga').last()
	status_obito_registro = status_registro.filter(descricao='Obito').last()
	status_aguard_lista_espera_registro = status_registro.filter(descricao='Aguardando em Lista de Espera').last()
	status_regulado_registro = status_registro.filter(descricao='Regulado').last()
	status_nao_regulado_registro = status_registro.filter(descricao='Não Preenche Critérios').last()
	status_aguard_inf_registro = status_registro.filter(descricao='Aguardando Informações').last()

	data_regulacao = registro.data_regulacao
	data_regulacao_template = None
	hora = None
	minuto = None
	if data_regulacao:
		data_regulacao_template = data_regulacao.strftime('%d/%m/%Y')
		hora_ = data_regulacao.hour
		hora = hora_ - 3
		minuto = data_regulacao.minute


	responsavel_pelo_preenchimento = request.user

	pa_admissao = registro.pa_admissao
	p = pa_admissao.split("x")
	pad_1 = p[0]
	pad_2 = p[1]

	pa = registro.pa
	p = pa.split("x")
	p1 = p[0]
	p2 = p[1]

	#status_registro = Status.objects.filter(registro_covid=registro.id)

	status_list_descricao_ = []

	for s in status_registro:
		status_list_descricao_.append(s.descricao)

	status_list_descricao = []
	if len(status_list_descricao_) > 0:
		status_list_descricao = status_list_descricao_.pop()

	#encoded = base64.b64encode(registro.image_descricao_clinica).decode('ascii')

	#img_tc = base64.b64encode(registro.image_laudo_tc).decode('ascii')

	#img_rx = base64.b64encode(registro.image_laudo_rx).decode('ascii')

	if not registro.image_descricao_clinica and not registro.image_laudo_tc and not registro.image_laudo_rx:
		if not regulacao_registro_last:
			
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
		else:
			senha = regulacao_registro_last.senha
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'senha':senha, 'hora':hora, 
			'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
			

	elif not registro.image_laudo_tc and not registro.image_laudo_rx:
		encoded = base64.b64encode(registro.image_descricao_clinica).decode('ascii')
		if regulacao_registro_last:
			senha = regulacao_registro_last.senha
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
				'p2':p2, 'data_regulacao_template':data_regulacao_template, 
				'status_list_descricao':status_list_descricao, 'encoded':encoded, 
				'status_aguard_conf_vaga_registro':status_aguard_conf_vaga_registro, 
				'status_obito_registro':status_obito_registro, 
				'status_aguard_lista_espera_registro':status_aguard_lista_espera_registro, 
				'status_regulado_registro':status_regulado_registro, 
				'status_nao_regulado_registro':status_nao_regulado_registro, 
				'status_aguard_inf_registro':status_aguard_inf_registro, 
				'senha':senha, 'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
		else:
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
				'p2':p2, 'data_regulacao_template':data_regulacao_template, 
				'status_list_descricao':status_list_descricao, 'encoded':encoded, 
				'status_aguard_conf_vaga_registro':status_aguard_conf_vaga_registro, 
				'status_obito_registro':status_obito_registro, 
				'status_aguard_lista_espera_registro':status_aguard_lista_espera_registro, 
				'status_regulado_registro':status_regulado_registro, 
				'status_nao_regulado_registro':status_nao_regulado_registro, 
				'status_aguard_inf_registro':status_aguard_inf_registro, 
				'hora':hora, 'minuto':minuto, 'regulacao_registro_last':
				regulacao_registro_last, 'pareceristas':pareceristas, 
				'pad_1':pad_1, 'pad_2':pad_2})
	elif not registro.image_descricao_clinica and not registro.image_laudo_tc:
		img_rx = base64.b64encode(registro.image_laudo_rx).decode('ascii')
		if regulacao_registro_last:
			senha = regulacao_registro_last.senha
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'senha':senha, 'img_rx':img_rx, 
			'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
		else:
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'img_rx':img_rx, 'hora':hora, 
			'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
	elif not registro.image_descricao_clinica and not registro.image_laudo_rx:
		img_tc = base64.b64encode(registro.image_laudo_tc).decode('ascii')
		if regulacao_registro_last:
			senha = regulacao_registro_last.senha
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'senha':senha, 'img_tc':img_tc, 
			'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
		else:
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'img_tc':img_tc, 'hora':hora, 
			'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
	elif not registro.image_laudo_rx:
		img_tc = base64.b64encode(registro.image_laudo_tc).decode('ascii')
		encoded = base64.b64encode(registro.image_descricao_clinica).decode('ascii')
		if regulacao_registro_last:
			senha = regulacao_registro_last.senha
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'senha':senha, 'img_tc':img_tc,
			 'encoded':encoded, 'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
		else:
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'img_tc':img_tc, 'encoded':encoded, 
			'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
	elif not registro.image_laudo_tc:
		img_rx = base64.b64encode(registro.image_laudo_rx).decode('ascii')
		encoded = base64.b64encode(registro.image_descricao_clinica).decode('ascii')
		if regulacao_registro_last:
			senha = regulacao_registro_last.senha
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'senha':senha, 'img_rx':img_rx,
			 'encoded':encoded, 'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
		else:
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'img_rx':img_rx, 'encoded':encoded, 
			'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})

	elif not registro.image_descricao_clinica and (registro.image_laudo_rx or registro.image_laudo_tc):
		img_tc = base64.b64encode(registro.image_laudo_tc).decode('ascii')
		img_rx = base64.b64encode(registro.image_laudo_rx).decode('ascii')
		if regulacao_registro_last:
			senha = regulacao_registro_last.senha
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'senha':senha, 'img_tc':img_tc, 'img_rx':
			img_rx, 'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
		else:
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'img_tc':img_tc, 'img_rx':
			img_rx, 'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})

	else:
		img_tc = base64.b64encode(registro.image_laudo_tc).decode('ascii')
		img_rx = base64.b64encode(registro.image_laudo_rx).decode('ascii')
		encoded = base64.b64encode(registro.image_descricao_clinica).decode('ascii')
		if regulacao_registro_last:
			senha = regulacao_registro_last.senha
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'senha':senha, 'img_tc':img_tc, 'img_rx':
			img_rx, 'encoded':encoded, 'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})
		else:
			return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
			'p2':p2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_regulado_registro':
			status_regulado_registro, 'status_nao_regulado_registro':
			status_nao_regulado_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'img_tc':img_tc, 'img_rx':
			img_rx, 'encoded':encoded, 'hora':hora, 'minuto':minuto, 
			'regulacao_registro_last':regulacao_registro_last, 'pareceristas':
			pareceristas, 'pad_1':pad_1, 'pad_2':pad_2})



	#return render(request, 'regulacao_detail.html', {'registro':registro, 'p1':p1, 
	#	'p2':p2, 'data_regulacao_template':data_regulacao_template, 
	#	'status_list_descricao':status_list_descricao, 'encoded':encoded, 
	#	'img_tc':img_tc, 'img_rx':img_rx, 'status_aguard_conf_vaga_registro':
	#	status_aguard_conf_vaga_registro, 'status_obito_registro':
	#	status_obito_registro, 'status_aguard_lista_espera_registro':
	#	status_aguard_lista_espera_registro, 'status_regulado_registro':
	#	status_regulado_registro, 'status_nao_regulado_registro':
	#	status_nao_regulado_registro, 'status_aguard_inf_registro':
	#	status_aguard_inf_registro, 'senha':senha})

@login_required
def regulacao_edit(request, id):

	#registro_id = request.GET.get('id')
	#registro = RegistroCovid.objects.get(id=registro_id)
	registro = RegistroCovid.objects.get(id=id)
	pa = registro.pa
	p = pa.split("x")
	p1 = p[0]
	p2 = p[1]

	municipios = Cnes.objects.order_by('MUNICIPIO').distinct('MUNICIPIO')

	return render(request, 'regulacao_edit.html', {'registro':registro, 
		'p1': p1, 'p2': p2, 'municipios':municipios})

@login_required
def regulacao_edit_set(request, id):

	registro = RegistroCovid.objects.get(id=id)

	
	responsavel_pelo_preenchimento = request.user

	#data = datetime.now()
	#data_notificacao = data.strftime("%d-%m-%Y")
	#hora_notificacao = data.strftime("%H:%M")

	#data_notificacao = request.POST.get('data_notificacao')
	#hora_notificacao = request.POST.get('hora_notificacao')
	
	nome_solicitante = request.POST.get('nome_solicitante')
	municipio_estabelecimento_solicitante = request.POST.get('municipio_do_estabelecimento')
	estabelecimento_solicitante = request.POST.get('estabelecimento_solicitante')
	estabelecimento_solicitante_outro = request.POST.get('desc_outro_estabelecimento')
	nome_paciente = request.POST.get('nome_paciente')
	
	idade_paciente_cap = request.POST.get('idade_paciente')
	if idade_paciente_cap != '':
		idade_paciente = int(idade_paciente_cap)
	else:
		idade_paciente = None
	
	sexo_paciente = request.POST.get('sexo_paciente')

	recurso_que_precisa = request.POST.get('recurso_que_precisa')
	#estado_origem = request.POST.get('estado_paciente')
	cidade_origem = request.POST.get('cidade_origem')
	telefone_retorno = request.POST.get('telefone_retorno')

	frequencia_respiratoria_cap = request.POST.get('frequencia_respiratoria')
	if frequencia_respiratoria_cap != '':
		frequencia_respiratoria = int(frequencia_respiratoria_cap)
	else:
		frequencia_respiratoria = None

	saturacao_paciente_cap = request.POST.get('saturacao_paciente')
	if saturacao_paciente_cap != '':
		saturacao_paciente = int(saturacao_paciente_cap)
	else:
		saturacao_paciente = None

	ar_o2 = request.POST.get('ar_o2')
	
	frequencia_cardiaca_cap = request.POST.get('f_cardiaca_paciente')
	if frequencia_cardiaca_cap != '':
		frequencia_cardiaca = int(frequencia_cardiaca_cap)
	else:
		frequencia_cardiaca = None

	pa_part1 = request.POST.get('pa_part1')
	pa_part2 = request.POST.get('pa_part2')

	pa_1 = pa_part1
	pa_2 = pa_part2

	pa = pa_part1 +"x"+ pa_part2
	conciencia = request.POST.get('consciencia_paciente')
	
	temperatura_cap = request.POST.get('temperatura_paciente')
	if temperatura_cap != '':
		temperatura = float(temperatura_cap)
	else:
		temperatura = None

	observacao = request.POST.get('observacao_paciente')

	
	cidade = request.POST.get('municipio_referencia')

	request.session['nome_solicitante'] = nome_solicitante
	request.session['idade_paciente'] = idade_paciente
	request.session['nome_paciente'] = nome_paciente
	request.session['sexo_paciente'] = sexo_paciente
	municipio_estabelecimento_solicitante = request.POST.get('municipio_estabelecimento_solicitante')
	request.session['estabelecimento_solicitante'] = estabelecimento_solicitante
	request.session['estabelecimento_solicitante_outro'] = estabelecimento_solicitante_outro
	request.session['recurso_que_precisa'] = recurso_que_precisa
	request.session['cidade_origem'] = cidade_origem
	request.session['telefone_retorno'] = telefone_retorno
	request.session['frequencia_respiratoria'] = frequencia_respiratoria
	request.session['saturacao_paciente'] = saturacao_paciente
	request.session['frequencia_cardiaca'] = frequencia_cardiaca
	request.session['conciencia'] = conciencia
	request.session['temperatura'] = temperatura
	request.session['observacao'] = observacao
	request.session['pa_1'] = pa_1
	request.session['pa_2'] = pa_2
	request.session['municipio_referencia'] = cidade

	return redirect('regulacao', id=id)

@login_required
def search_register(request):
	registros = RegistroCovid.objects.all()

	#paginator = Paginator(registros, 5)
	#page = request.GET.get('page')
	#regs = paginator.get_page(page)

	return render(request, 'search_register.html', {'registros':registros})

@login_required
def search_between_date(request):

	return render(request, 'search_between_date.html')

@login_required
def search_between_date_set(request):
	data_inicio = request.POST.get('data_inicio')
	data_fim = request.POST.get('data_fim')

	values = RegistroCovid.objects.filter(data_notificacao__range=[data_inicio, data_fim])
		
	return render(request, 'result_search_between_date.html', {'values':values})

@login_required
def result_search_between_date(request):


	return render(request, 'result_search_between_date.html')


@login_required
def gerar_relatorios(request):
	
	return render(request, 'gerar_relatorios.html')


@login_required
def gerar_relatorios_set(request):
	data_inicio = request.POST.get('data_inicio')
	data_fim = request.POST.get('data_fim')


	values = RegistroCovid.objects.filter(data_notificacao__range=[data_inicio, data_fim])
	#values = RegistroCovid.objects.all()
	status = Status.objects.all()
	
	for s in status:
		for v in values:
			if v.id == s.registro_covid_id and s.descricao != None:
				regulacao_ = s.descricao
				#print(regulacao_)
				
		

	#for vl in values:
	#	print(', '.join(vl.comorbidades))

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

	response['Content-Disposition'] = 'attachment; filename=myfile.xlsx'

	workbook = Workbook()
	worksheet = workbook.active

	columns = [
		'Senha',
		'Registro CERN',
		'Nome',
		'Idade',
		'Sexo',
		'Regulacao',
		'Município de Residencia',
		'Estabelecimento de Origem',
		'Estabelecimento Referencia Covid',
		'Município do Estabelecimento',
		'Perfil',
		'Data',
		'Comorbidades',
		'Nível de Atenção',
		'Tipo de Teste',
		'Teste Covid',

	]

	row_num = 1

	regulacao_status = []


	for col_num, column_title in enumerate(columns, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	
	for v in values:
		result_exames = []
		row_num += 1
		

		#comorbidades_list = v.comorbidades
		#comorbidades = str(comorbidades_list)
		#v.comorbidades.reverse()
		#print(v.comorbidades)
		#if v.regulacao_paciente == 'Paciente preenche critérios para Regulação':
		#	v.regulacao_paciente = 'Regulado'

		#if v.regulacao_paciente == 'Paciente não preenche critérios para Regulação':
		#	v.regulacao_paciente = 'Não Regulado'
		sexo = ""
		if v.sexo_paciente != None:
			sexo = v.sexo_paciente

		estabelecimento_referencia_covid = ""
		if v.estabelecimento_referencia != None:
			estabelecimento_referencia_covid = v.estabelecimento_referencia
		
		perfil = ""
		if v.leito != None:
			perfil = v.leito

		nivel_atencao = ""
		if v.news_fast_pb != None:
			nivel_atencao = v.news_fast_pb
		elif v.news_modificado != None:
			nivel_atencao = v.news_modificado
		#nivel_atencao_ = nivel_atencao

		if v.pesquisa_teste_sars_cov_2 != None:
			result_exames.append(v.pesquisa_teste_sars_cov_2)
		if v.rt_pcr_sars_cov_2 != None:
			result_exames.append(v.rt_pcr_sars_cov_2)

		result_exames_ = str(result_exames)[1 : -1]


		testes_covid = ""
		if v.pesquisa_teste_sars_cov_2 != None and v.rt_pcr_sars_cov_2 != None:
			testes_covid = "Teste rápido, PCR"
		elif v.pesquisa_teste_sars_cov_2 != None and v.rt_pcr_sars_cov_2 == None:
			testes_covid = "Teste rápido"
		elif v.pesquisa_teste_sars_cov_2 == None and v.rt_pcr_sars_cov_2 != None:
			testes_covid = "PCR"
		


		if v.idade_paciente != None or v.idade_paciente != '':
			idade = v.idade_paciente

		if v.estabelecimento_solicitante != None or v.estabelecimento_solicitante != '':
			estabelecimento_solicitante = v.estabelecimento_solicitante

		if v.data_regulacao:
			data_regulacao = v.data_regulacao.strftime("%d-%m-%Y")
		else:
			data_regulacao = None

		if v.comorbidades != '' or v.comorbidades != None:
			comorbidades = v.comorbidades

		if v.municipio_estabelecimento_referencia != None or v.municipio_estabelecimento_referencia != '':
			municipio_estabelecimento_referencia = v.municipio_estabelecimento_referencia

		if v.codigo_sescovid != '' or v.codigo_sescovid != None:
			codigo_sescovid = v.codigo_sescovid

		if v.last_status != None:
			last_status = v.last_status


		#regulacao = ""
		#for s in status:
		#	if v.id == s.registro_covid_id and s.descricao != None:
		#		v.regulacao_status = s.descricao

		#	if v.regulacao_status != None:
		#		regulacao = v.regulacao_status
				#print(v.regulacao_status)


		row = [
			codigo_sescovid,
			v.codigo_registro_completo,
			v.nome_paciente,
			idade,
			sexo,
			last_status,
			v.municipio_estabelecimento_solicitante,
			estabelecimento_solicitante,
			estabelecimento_referencia_covid,
			municipio_estabelecimento_referencia,
			perfil,
			data_regulacao,
			comorbidades,
			nivel_atencao,
			testes_covid,
			result_exames_,
		]
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell_value_ = str(cell_value)
			cell.value = cell_value_

	workbook.save(response)




	return response

	#return render(request, 'result_for_relatorios.html', {'values':values})

@login_required
def result_for_relatorios(request):
	return render(request, 'result_for_relatorios.html')


@login_required
def status_registro(request, id):
	registro = RegistroCovid.objects.get(id=id)
	
	status_registro = Status.objects.filter(registro_covid=registro.id)
	status_registro_last = Status.objects.filter(registro_covid=registro.id).last()


	#for f in status_registro:
	#	print(f,"data:",f.data_notificacao,"hora",f.hora_notificacao)

	#status_list_descricao = []

	#for s in status_registro:
	#	status_list_descricao.append(s.descricao)


	status_aguard_conf_vaga_registro = status_registro.filter(descricao='Aguardando confirmação de Vaga').last()
	status_obito_registro = status_registro.filter(descricao='Obito').last()
	status_aguard_lista_espera_registro = status_registro.filter(descricao='Aguardando em Lista de Espera').last()
	status_regulado_registro = status_registro.filter(descricao='Regulado').last()
	status_nao_regulado_registro = status_registro.filter(descricao='Não Preenche Critérios').last()
	status_aguard_inf_registro = status_registro.filter(descricao='Aguardando Informações').last()

	data_cancelamento = datetime.now()
	formateDate = data_cancelamento.strftime("%d-%m-%Y")
	hora = data_cancelamento.strftime("%H:%M")

	hora_canc = 0
	minuto_canc = 0

	if registro.data_cancelamento:
		hora_canc_ = registro.data_cancelamento.hour
		minuto_canc = registro.data_cancelamento.minute
		hora_canc = hora_canc_ - 3


	data_cancelamento_bd = registro.data_cancelamento
	if data_cancelamento_bd:
		data_cancelamento_bd_template = data_cancelamento_bd.strftime("%d-%m-%Y")
		#hora_cancelamento_bd = data_cancelamento_bd.strftime("%H:%M")
		#hora_cancelamento_bd_template = hora_cancelamento_bd - timedelta(hours=3)
		if len(str(minuto_canc)) < 2:
			pref = 0
			return render(request, 'status_registro.html', {'registro':registro, 
				'status_registro':status_registro, 
				'formateDate':formateDate, 'hora':hora, 'status_registro_last':
				status_registro_last, 'status_regulado_registro':
				status_regulado_registro, 'status_nao_regulado_registro':
				status_nao_regulado_registro, 
				'status_aguard_lista_espera_registro':
				status_aguard_lista_espera_registro, 
				'status_obito_registro':status_obito_registro, 
				'status_aguard_conf_vaga_registro':
				status_aguard_conf_vaga_registro, 'status_aguard_inf_registro':
				status_aguard_inf_registro, 'data_cancelamento_bd_template':
				data_cancelamento_bd_template, 'hora_canc':hora_canc, 
				'minuto_canc':minuto_canc, 'pref':pref})
		else:
			return render(request, 'status_registro.html', {'registro':registro, 
				'status_registro':status_registro, 
				'formateDate':formateDate, 'hora':hora, 'status_registro_last':
				status_registro_last, 'status_regulado_registro':
				status_regulado_registro, 'status_nao_regulado_registro':
				status_nao_regulado_registro, 
				'status_aguard_lista_espera_registro':
				status_aguard_lista_espera_registro, 
				'status_obito_registro':status_obito_registro, 
				'status_aguard_conf_vaga_registro':
				status_aguard_conf_vaga_registro, 'status_aguard_inf_registro':
				status_aguard_inf_registro, 'data_cancelamento_bd_template':
				data_cancelamento_bd_template, 'hora_canc':hora_canc, 
				'minuto_canc':minuto_canc})
	else:
		if len(str(minuto_canc)) < 2:
			pref = 0
			return render(request, 'status_registro.html', {'registro':registro,
			'status_registro':status_registro, 'formateDate':formateDate, 
			'hora':hora, 'status_registro_last':status_registro_last, 
			'status_regulado_registro':status_regulado_registro, 
			'status_nao_regulado_registro':status_nao_regulado_registro, 
			'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'hora_canc':hora_canc, 'minuto_canc':
			minuto_canc, 'pref':pref})
		else:
			return render(request, 'status_registro.html', {'registro':registro,
			'status_registro':status_registro, 'formateDate':formateDate, 
			'hora':hora, 'status_registro_last':status_registro_last, 
			'status_regulado_registro':status_regulado_registro, 
			'status_nao_regulado_registro':status_nao_regulado_registro, 
			'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 'status_obito_registro':
			status_obito_registro, 'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'status_aguard_inf_registro':
			status_aguard_inf_registro, 'hora_canc':hora_canc, 'minuto_canc':
			minuto_canc})


@login_required
def status_registro_set(request, id):
	registro = RegistroCovid.objects.get(id=id)

	andamento_processo = ''
	justificativa_cancelamento = request.POST.get('justif_cancelar_regulacao')
	if justificativa_cancelamento != '':
		andamento_processo = "Cancelado"

	data_cancelamento = datetime.now()
	#hora_cancelamento = request.POST.get('')

	registro.andamento_processo = andamento_processo
	registro.justificativa_cancelamento = justificativa_cancelamento
	registro.data_cancelamento = data_cancelamento
	#registro.hora_cancelamento = hora_cancelamento
	registro.save()

	return redirect('status_registro', id=id)

@login_required
def remove_registro_covid(request, id):
	registro = RegistroCovid.objects.filter(id=id).delete()

	return redirect('/')


@login_required
def image_descricao_clinica_alter(request, id):
	registro = RegistroCovid.objects.get(id=id)
	if registro.image_descricao_clinica:
		encoded = base64.b64encode(registro.image_descricao_clinica).decode('ascii')
		return render(request, 'image_descricao_clinica_alter.html', {'registro':
		registro, 'encoded':encoded})
	else:
		return render(request, 'image_descricao_clinica_alter.html', {'registro':
		registro})


@login_required
def image_descricao_clinica_alter_set(request, id):
	registro = RegistroCovid.objects.get(id=id)

	image_descricao_clinica = request.FILES['file_descricao_clinica'].file.read()
	registro.image_descricao_clinica = image_descricao_clinica
	registro.save()
	return redirect('regulacao_detail', id=id)


@login_required
def image_rx_torax_alter(request, id):
	registro = RegistroCovid.objects.get(id=id)

	if registro.image_laudo_rx:
		img_rx = base64.b64encode(registro.image_laudo_rx).decode('ascii')
		return render(request, 'image_rx_torax_alter.html', {'registro':
		registro, 'img_rx':img_rx})
	else:
		return render(request, 'image_rx_torax_alter.html', {'registro':
		registro})


@login_required
def image_rx_torax_alter_set(request, id):
	registro = RegistroCovid.objects.get(id=id)

	image_laudo_rx = request.FILES['file_rx_torax'].file.read()
	registro.image_laudo_rx = image_laudo_rx
	registro.save()
	return redirect('regulacao_detail', id=id)


@login_required
def image_tc_torax_alter(request, id):
	registro = RegistroCovid.objects.get(id=id)

	if registro.image_laudo_tc:
		img_tc = base64.b64encode(registro.image_laudo_tc).decode('ascii')
		return render(request, 'image_tc_torax_alter.html', {'registro':
		registro, 'img_tc':img_tc})
	else:
		return render(request, 'image_tc_torax_alter.html', {'registro':
		registro})


@login_required
def image_tc_torax_alter_set(request, id):
	registro = RegistroCovid.objects.get(id=id)

	image_laudo_tc = request.FILES['file_tc_torax'].file.read()
	registro.image_laudo_tc = image_laudo_tc
	registro.save()
	return redirect('regulacao_detail', id=id)


@login_required
def regular_registro(request, id):
	registro = RegistroCovid.objects.get(id=id)

	regulacoes_registro = Regulacao.objects.filter(registro_covid_id=registro.id)

	regulacao_registro_last = regulacoes_registro.last()

	status_registro = Status.objects.filter(registro_covid=registro.id)
	status_registro_last = Status.objects.filter(registro_covid=registro.id).last()

	status_aguard_conf_vaga_registro = status_registro.filter(descricao='Aguardando confirmação de Vaga').last()
	status_obito_registro = status_registro.filter(descricao='Obito').last()
	status_aguard_lista_espera_registro = status_registro.filter(descricao='Aguardando em Lista de Espera').last()
	status_regulado_registro = status_registro.filter(descricao='Regulado').last()
	status_nao_regulado_registro = status_registro.filter(descricao='Não Preenche Critérios').last()
	status_aguard_inf_registro = status_registro.filter(descricao='Aguardando Informações').last()


	data_cancelamento = registro.data_cancelamento
	hora = None
	minuto = None
	if data_cancelamento:
		data_cancelamento_template = data_cancelamento.strftime('%d/%m/%Y')
		hora_ = data_cancelamento.hour
		minuto = data_cancelamento.minute
		hora = hora_ - 3
		return render(request, 'regular_registro.html', {'registro':registro, 
		'regulacoes_registro':regulacoes_registro, 'hora':hora, 'minuto':minuto, 
		'data_cancelamento_template':data_cancelamento_template, 
		'status_aguard_conf_vaga_registro':status_aguard_conf_vaga_registro, 
		'status_obito_registro':status_obito_registro, 
		'status_aguard_lista_espera_registro':status_aguard_lista_espera_registro,
		'status_regulado_registro':status_regulado_registro, 
		'status_nao_regulado_registro':status_nao_regulado_registro, 
		'status_aguard_inf_registro':status_aguard_inf_registro, 
		'status_registro_last':status_registro_last})
	else:
		return render(request, 'regular_registro.html', {'registro':registro, 
		'regulacoes_registro':regulacoes_registro, 
		'status_aguard_conf_vaga_registro':status_aguard_conf_vaga_registro, 
		'status_obito_registro':status_obito_registro, 
		'status_aguard_lista_espera_registro':status_aguard_lista_espera_registro,
		'status_regulado_registro':status_regulado_registro, 
		'status_nao_regulado_registro':status_nao_regulado_registro, 
		'status_aguard_inf_registro':status_aguard_inf_registro, 
		'status_registro_last':status_registro_last})


@login_required
def regular_registro_set(request, id):
	registro = RegistroCovid.objects.get(id=id)

	if registro.senha != '' or registro.senha != None:
		senha_bd = registro.senha
	else:
		senha_bd = 0


	return redirect('regulacao_detail', id=id)


@login_required
def paciente_atribuir_senha(request, id):
	data_regulacao = datetime.now()
	data_regulacao_template = data_regulacao.strftime('%d/%m/%Y %H:%M')

	registro = RegistroCovid.objects.get(id=id)

	pareceristas_bd = registro.pareceristas
	pareceristas = []
	if pareceristas_bd:
		pareceristas = list(set(pareceristas_bd))

	regulacoes_registro = Regulacao.objects.filter(registro_covid_id=registro.id)

	#Ultima regulacao
	regulacao_last = Regulacao.objects.all().last()
	#Ultima regulacao do paciente:
	regulacao_paciente_last = regulacoes_registro.last()
	if regulacao_last:
		senha_last = regulacao_last.senha
		senha_new = senha_last + 1
	else:
		senha_new = 500000


	senha = senha_new

	codigo_sescovid = "SESCOVID" + str(senha_new)

	pa = registro.pa

	p = pa.split("x")
	pa_1 = p[0]
	pa_2 = p[1]

	status_registro = Status.objects.filter(registro_covid=registro.id)

	status_list_descricao_ = []

	for s in status_registro:
		status_list_descricao_.append(s.descricao)



	status_list_descricao = []
	if len(status_list_descricao_) != 0:
		status_list_descricao = status_list_descricao_.pop()
	else:
		status_list_descricao
	status_aguard_conf_vaga_registro = Status.objects.filter(descricao='Aguardando confirmação de Vaga').last()
	#status_obito_registro = Status.objects.filter(descricao='Obito').last()
	status_aguard_lista_espera_registro = Status.objects.filter(descricao='Aguardando em Lista de Espera').last()
	status_regulado_registro = Status.objects.filter(descricao='Regulado').last()
	
	#status_nao_regulado_registro = []
	status_nao_regulado_registro = Status.objects.filter(descricao='Não Regulado').last()
	#status_nao_regulado = status_nao_regulado_registro.last()


	municipio_estabelecimento_referencia = request.POST.get('municipio_estabelecimento_referencia')
	#request.session['nome_solicitante'] = nome_solicitante
	#request.session['idade_paciente'] = idade_paciente
	#request.session['nome_paciente'] = nome_paciente
	#request.session['sexo_paciente'] = sexo_paciente
	municipio_estabelecimento_solicitante = request.POST.get('municipio_estabelecimento_solicitante')
	estabelecimento_solicitante = request.POST.get('estabelecimento_solicitante')
	estabelecimento_solicitante_outro = request.POST.get('estabelecimento_solicitante_outro')
	#request.session['recurso_que_precisa'] = recurso_que_precisa
	#request.session['cidade_origem'] = cidade_origem
	#request.session['telefone_retorno'] = telefone_retorno
	#request.session['frequencia_respiratoria'] = frequencia_respiratoria
	#request.session['saturacao_paciente'] = saturacao_paciente
	#request.session['frequencia_cardiaca'] = frequencia_cardiaca
	#request.session['conciencia'] = conciencia
	#request.session['temperatura'] = temperatura
	#request.session['observacao'] = observacao
	#request.session['pa_1'] = pa_1
	#request.session['pa_2'] = pa_2
	cidade = request.POST.get('municipio_referencia')


	estabelecimentos = Cnes.objects.filter(MUNICIPIO = municipio_estabelecimento_referencia).order_by('NO_FANTASIA')

	encoded_ = None
	if registro.image_descricao_clinica:
		encoded_ = base64.b64encode(registro.image_descricao_clinica).decode('ascii')

	encoded = encoded_

	if not registro.image_laudo_tc and not registro.image_laudo_rx:
		return render(request, 'paciente_atribuir_senha.html', {'registro' : registro, 
			'codigo_sescovid':codigo_sescovid, 'senha':senha, 'pa_1':pa_1, 
			'pa_2':pa_2, 
			'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro, 'pareceristas':pareceristas,
			'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 
			'status_regulado_registro':status_regulado_registro, 
			'status_nao_regulado_registro':status_nao_regulado_registro, 
			'municipio_estabelecimento_solicitante':
			municipio_estabelecimento_solicitante, 
			'estabelecimento_solicitante':estabelecimento_solicitante, 
			'estabelecimento_solicitante_outro':
			estabelecimento_solicitante_outro, 'cidade':cidade,
			'estabelecimentos':estabelecimentos,
			'municipio_estabelecimento_referencia':
			municipio_estabelecimento_referencia, 'encoded':encoded})
	elif not registro.image_laudo_tc and not registro.image_laudo_rx and not registro.image_descricao_clinica:
		return render(request, 'paciente_atribuir_senha.html', {'registro' : registro, 
			'codigo_sescovid':codigo_sescovid, 'senha':senha, 'pa_1':pa_1, 
			'pa_2':pa_2, 'data_regulacao_template':data_regulacao_template, 
			'status_list_descricao':status_list_descricao, 
			'status_aguard_conf_vaga_registro':
			status_aguard_conf_vaga_registro,
			'status_aguard_lista_espera_registro':
			status_aguard_lista_espera_registro, 
			'status_regulado_registro':status_regulado_registro, 
			'status_nao_regulado_registro':status_nao_regulado_registro, 
			'municipio_estabelecimento_solicitante':
			municipio_estabelecimento_solicitante, 
			'estabelecimento_solicitante':estabelecimento_solicitante, 
			'estabelecimento_solicitante_outro':
			estabelecimento_solicitante_outro, 'cidade':cidade,
			'estabelecimentos':estabelecimentos,
			'municipio_estabelecimento_referencia':
			municipio_estabelecimento_referencia, 'pareceristas':pareceristas})

	elif not registro.image_laudo_rx:
		img_tc = base64.b64encode(registro.image_laudo_tc).decode('ascii')

		return render(request, 'paciente_atribuir_senha.html', {'registro' : registro, 
		'codigo_sescovid':codigo_sescovid, 'senha':senha, 'pa_1':pa_1, 
		'pa_2':pa_2, 'data_regulacao_template': data_regulacao_template,
		'status_list_descricao':status_list_descricao, 'status_aguard_conf_vaga_registro':
		status_aguard_conf_vaga_registro, 'status_aguard_lista_espera_registro':
		status_aguard_lista_espera_registro, 'status_regulado_registro':
		status_regulado_registro, 'status_nao_regulado_registro':
		status_nao_regulado_registro, 'municipio_estabelecimento_solicitante':
		municipio_estabelecimento_solicitante, 'estabelecimento_solicitante':
		estabelecimento_solicitante, 'estabelecimento_solicitante_outro':
		estabelecimento_solicitante_outro, 'cidade':cidade, 'estabelecimentos':
		estabelecimentos, 'municipio_estabelecimento_referencia':
		municipio_estabelecimento_referencia, 'encoded':encoded, 'img_tc':
		img_tc, 'pareceristas':pareceristas})
	elif not registro.image_laudo_tc:
		img_rx = base64.b64encode(registro.image_laudo_rx).decode('ascii')

		return render(request, 'paciente_atribuir_senha.html', {'registro' : registro, 
		'codigo_sescovid':codigo_sescovid, 'senha':senha, 'pa_1':pa_1, 
		'pa_2':pa_2, 'data_regulacao_template': data_regulacao_template,
		'status_list_descricao':status_list_descricao, 'status_aguard_conf_vaga_registro':
		status_aguard_conf_vaga_registro, 'status_aguard_lista_espera_registro':
		status_aguard_lista_espera_registro, 'status_regulado_registro':
		status_regulado_registro, 'status_nao_regulado_registro':
		status_nao_regulado_registro, 'municipio_estabelecimento_solicitante':
		municipio_estabelecimento_solicitante, 'estabelecimento_solicitante':
		estabelecimento_solicitante, 'estabelecimento_solicitante_outro':
		estabelecimento_solicitante_outro, 'cidade':cidade, 'estabelecimentos':
		estabelecimentos, 'municipio_estabelecimento_referencia':
		municipio_estabelecimento_referencia, 'encoded':encoded, 'img_rx':img_rx, 
		'pareceristas':pareceristas})
	else:
		img_tc = base64.b64encode(registro.image_laudo_tc).decode('ascii')
		img_rx = base64.b64encode(registro.image_laudo_rx).decode('ascii')

		return render(request, 'paciente_atribuir_senha.html', {'registro' : registro,
		'codigo_sescovid':codigo_sescovid, 'senha':senha, 'pa_1':pa_1, 'pa_2':pa_2, 
		'data_regulacao_template': data_regulacao_template,
		'status_list_descricao':status_list_descricao, 'status_aguard_conf_vaga_registro':
		status_aguard_conf_vaga_registro, 'status_aguard_lista_espera_registro':
		status_aguard_lista_espera_registro, 'status_regulado_registro':
		status_regulado_registro, 'status_nao_regulado_registro':
		status_nao_regulado_registro, 'municipio_estabelecimento_solicitante':
		municipio_estabelecimento_solicitante, 'estabelecimento_solicitante':
		estabelecimento_solicitante, 'estabelecimento_solicitante_outro':
		estabelecimento_solicitante_outro, 'cidade':cidade, 'estabelecimentos':
		estabelecimentos, 'municipio_estabelecimento_referencia':
		municipio_estabelecimento_referencia, 'encoded':encoded, 'img_tc':
		img_tc, 'img_rx':img_rx, 'pareceristas':pareceristas})


@login_required
def paciente_atribuir_senha_set(request, id):
	registro = RegistroCovid.objects.get(id=id)

	responsavel_pelo_preenchimento = request.user
	
	
	nome_solicitante = request.POST.get('nome_solicitante')
	municipio_estabelecimento_solicitante = registro.municipio_estabelecimento_solicitante
	estabelecimento_solicitante = registro.estabelecimento_solicitante
	estabelecimento_solicitante_outro = registro.estabelecimento_solicitante_outro
	
	if estabelecimento_solicitante == '' or estabelecimento_solicitante == None:
		estabelecimento_solicitante = request.POST.get('estabelecimento_solicitante')
	
	if estabelecimento_solicitante_outro == '' or estabelecimento_solicitante_outro == None:
		estabelecimento_solicitante_outro = request.POST.get('estabelecimento_solicitante_outro')

	municipio_estabelecimento_referencia = request.POST.get('municipio_estabelecimento_referencia')
	estabelecimento_referencia = request.POST.get('estabelecimento_referencia')
	estabelecimento_referencia_outro = request.POST.get('estabelecimento_referencia_outro')
	nome_paciente = request.POST.get('nome_paciente')
	
	idade_paciente_cap = request.POST.get('idade_paciente')
	if idade_paciente_cap == '' or idade_paciente_cap == None:
		idade_paciente = None
	else:
		idade_paciente = int(float(idade_paciente_cap))

	sexo_paciente = request.POST.get('sexo_paciente')
	recurso_que_precisa = registro.recurso_que_precisa
	estado_origem = registro.estado_origem
	cidade_origem = registro.cidade_origem
	telefone_retorno = registro.telefone_retorno
	frequencia_respiratoria = registro.frequencia_respiratoria
	saturacao_paciente = registro.saturacao_paciente
	ar_o2 = registro.ar_o2
	frequencia_cardiaca = registro.frequencia_cardiaca
	
	pa_part1 = request.POST.get('pa_part1')
	pa_part2 = request.POST.get('pa_part2')
	pa = pa_part1 + "x" + pa_part2
	
	conciencia = registro.conciencia
	temperatura = registro.temperatura

	descricao_clinica = request.POST.get('descricao_clinica')

	#image_descricao_clinica = request.FILES['file_descricao_clinica'].file.read()

	#image_laudo_tc = request.FILES['file_tc_torax_regulacao'].file.read()
	#image_laudo_rx = request.FILES['file_rx_torax_regulacao'].file.read()
	
	sindrome_gripal = request.POST.getlist('s_gripal')
	
	tempo_quadro_sintomatico_cap = request.POST.get('tempo_inicio_sintomas')
	if tempo_quadro_sintomatico_cap == '' or tempo_quadro_sintomatico_cap == None:
		tempo_quadro_sintomatico = None
	else:
		tempo_quadro_sintomatico = int(float(tempo_quadro_sintomatico_cap))

	exposicao_pessoa_infectada = request.POST.get('exposicao_epidemiologica_sim')
	parentesco = request.POST.get('grau_parentesco')
	comorbidades = request.POST.getlist('comorbidades')
	outras_comorbidades = request.POST.get('descricao_outras_comorbidades')
	outras_cardiopatias = request.POST.get('desc_outras_cardiopatias')
	idade_gestacional = request.POST.get('idade_gestacao')
	comorbidades_obstetricas = request.POST.get('comorbidade_obstetrica')
	gesta_para = request.POST.get('gestacao_para')
	medicamentos_uso_habitual = request.POST.getlist('medicamento_uso_habitual')
	medicamento_outros = request.POST.get('descricao_outros_medicamento')
	
	spo2_cap = request.POST.get('spo2')
	if spo2_cap == '' or spo2_cap == None:
		spo2 = None
	else:
		spo2 = int(float(spo2_cap))
		

	fr_irpm_cap = request.POST.get('fr_irpm')
	if fr_irpm_cap == '' or fr_irpm_cap == None:
		fr_irpm = None
	else:
		fr_irpm = int(float(fr_irpm_cap))
		

	ventilacao_tipo = request.POST.get('ventilacao')
	o2_suporte = request.POST.getlist('o2_suporte')
	
	dose_cn_cap = request.POST.get('dose_cn')
	if dose_cn_cap == '' or dose_cn_cap == None:
		dose_cn = None
	else:
		dose_cn = float(dose_cn_cap)

	dose_venturi_cap = request.POST.get('dose_venturi')
	if dose_venturi_cap == '' or dose_venturi_cap == None:
		dose_venturi = None
	else:
		dose_venturi = float(dose_venturi_cap)
	

	vmi = request.POST.getlist('vmi')
	
	vt_cap = request.POST.get('vt')
	if vt_cap == '' or vt_cap == None:
		vt = None
	else:
		vt = float(vt_cap)
		

	delta_pressure_cap = request.POST.get('delta_pressure')
	if delta_pressure_cap == '' or delta_pressure_cap == None:
		delta_pressure = None
	else:
		delta_pressure = float(delta_pressure_cap)
		

	pplato_cap = request.POST.get('pplato')
	if pplato_cap == '' or pplato_cap == None:
		pplato = None
	else:
		pplato = float(pplato_cap)
		

	raw_cap = request.POST.get('raw')
	if raw_cap == '' or raw_cap == None:
		raw = None
	else:
		raw = float(raw_cap)
		

	cest_cap = request.POST.get('cest')
	if cest_cap == '' or cest_cap == None :
		cest = None
	else:
		cest = float(cest_cap)
		

	sao2_cap = request.POST.get('sao2')
	if sao2_cap == '' or sao2_cap == None:
		sao2 = None
	else:
		sao2 = float(sao2_cap)

	pao2_cap = request.POST.get('pao2')
	if pao2_cap == '' or pao2_cap == None:
		pao2 = None
	else:
		pao2 = float(pao2_cap)

	fio2_cap = request.POST.get('fio2')
	if fio2_cap == '' or fio2_cap == None:
		fio2 = None
	else:
		fio2 = float(fio2_cap)

	paco2_cap = request.POST.get('paco2')
	if paco2_cap == '' or paco2_cap == None:
		paco2 = None
	else:
		paco2 = float(paco2_cap)
	
	fc_cap = request.POST.get('fc')
	if fc_cap == '' or fc_cap == None:
		fc = None
	else:
		fc = int(float(fc_cap))
		

	temperatura_axilar_cap = request.POST.get('temp_auxiliar')
	if temperatura_axilar_cap == '' or temperatura_axilar_cap == None :
		temperatura_axilar = None
	else:
		temperatura_axilar = float(temperatura_axilar_cap)
		

	droga_vasoativa = request.POST.getlist('droga_vasoativa')
	
	qtd_nora_cap = request.POST.get('qtd_nora')
	if qtd_nora_cap == '' or qtd_nora_cap == None:
		qtd_nora = None
	else:
		qtd_nora = float(qtd_nora_cap)

	qtd_adrenalina_cap = request.POST.get('qtd_adrenalina')
	if qtd_adrenalina_cap == '' or qtd_adrenalina_cap == None:
		qtd_adrenalina = None
	else:
		qtd_adrenalina = float(qtd_adrenalina_cap)

	qtd_vasopressina_cap = request.POST.get('qtd_vasopressina')
	if qtd_vasopressina_cap == '' or qtd_vasopressina_cap == None:
		qtd_vasopressina = None
	else:
		qtd_vasopressina = float(qtd_vasopressina_cap)

	qtd_dobutamina_cap = request.POST.get('qtd_dobutamina')
	if qtd_dobutamina_cap == '' or qtd_dobutamina_cap == None:
		qtd_dobutamina = None
	else:
		qtd_dobutamina = float(qtd_dobutamina_cap)

	arritmia = request.POST.getlist('arritmia')
	infeccao_bacteriana = request.POST.get('infeccao_bacteriana')
	foco = request.POST.getlist('foco')
	uso_antibioticoterapia = request.POST.get('uso_antibioticoterapia')
	pesquisa_teste_sars_cov_2 = request.POST.get('pesquisa_sars_cov2')
	igg_cap = request.POST.get('igg')
	igg_bd = request.POST.get('igg_bd')
	if igg_cap == '' or igg_cap == None:
		if igg_bd == '' or igg_bd == None:
			igg = None
		else:
			igg = igg_bd
	else:
		igg = igg_cap

	igm_do_bd = registro.igm
	igm_cap = request.POST.get('igm')

	igm_bd_cap = request.POST.get('igm_bd')
	if igm_cap == '' or igm_cap == None:
		if igm_bd_cap == '' or igm_bd_cap == None:
			igm = None
		else:
			igm = igm_bd_cap
	else:
		igm = igm_cap

	
	data_igm_igg_cap = request.POST.get('data_igm_igg')
	data_igm_igg_bd_cap = request.POST.get('data_igm_igg_bd')

	if data_igm_igg_cap == '' or data_igm_igg_cap == None:
		if data_igm_igg_bd_cap == '' or data_igm_igg_bd_cap == None:
			data_igm_igg = None
		else:
			data_igm_igg = datetime.strptime(data_igm_igg_bd_cap, '%d/%m/%Y').date()
	else:
		data_igm_igg = data_igm_igg_cap


	rt_pcr_sars_cov_2 = request.POST.get('pcr_sars_cov2')
	
	data_da_coleta_bd_cap = request.POST.get('data_da_coleta_bd')
	data_coleta_cap = request.POST.get('data_da_coleta')
	if data_coleta_cap == '' or data_coleta_cap == None:
		if data_da_coleta_bd_cap == '' or data_da_coleta_bd_cap == None:
			data_coleta = None
		else:
			data_coleta = datetime.strptime(data_da_coleta_bd_cap, '%d/%m/%Y').date()
	else:
		data_coleta = data_coleta_cap
	
	em_uso_corticosteroide = request.POST.get('em_uso_corticosteroide')
	dose_corticosteroide = request.POST.get('dose_corticosteroide')
	
	data_inicio_corticosteroide_bd_cap = request.POST.get('data_inicio_corticosteroide_bd')
	data_inicio_corticosteroide_cap = request.POST.get('data_inicio_corticosteroide')
	if data_inicio_corticosteroide_cap == '' or data_inicio_corticosteroide_cap == None:
		if data_inicio_corticosteroide_bd_cap == '' or data_inicio_corticosteroide_bd_cap == None:
			data_inicio_corticosteroide = None
		else:
			data_inicio_corticosteroide = datetime.strptime(data_inicio_corticosteroide_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_corticosteroide = data_inicio_corticosteroide_cap
	

	em_uso_hidroxicloroquina = request.POST.get('em_uso_hidrocloroquina')
	
	data_inicio_hidroxicloroquina_cap = request.POST.get('data_inicio_hidroxicloroquina')
	data_inicio_hidroxicloroquina_bd_cap = request.POST.get('data_inicio_hidroxicloroquina_bd')


	if data_inicio_hidroxicloroquina_cap == '' or data_inicio_hidroxicloroquina_cap == None:
		if data_inicio_hidroxicloroquina_bd_cap == '' or data_inicio_hidroxicloroquina_bd_cap == None:
			data_inicio_hidroxicloroquina = None	
		else:
			data_inicio_hidroxicloroquina = datetime.strptime(data_inicio_hidroxicloroquina_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_hidroxicloroquina = data_inicio_hidroxicloroquina_cap


	

	em_uso_oseltamivir = request.POST.get('em_uso_oseltamivir')
	
	data_inicio_oseltamivir_cap = request.POST.get('data_inicio_oseltamivir')
	data_inicio_oseltamivir_bd_cap = request.POST.get('data_inicio_oseltamivir_bd')


	if data_inicio_oseltamivir_cap == '' or data_inicio_oseltamivir_cap == None:
		if data_inicio_oseltamivir_bd_cap == '' or data_inicio_oseltamivir_bd_cap == None:
			data_inicio_oseltamivir = None
		else:
			data_inicio_oseltamivir = datetime.strptime(data_inicio_oseltamivir_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_oseltamivir = data_inicio_oseltamivir_cap



	em_uso_heparina = request.POST.get('em_uso_heparina')

	data_inicio_heparina_bd_cap = request.POST.get('data_inicio_heparina_bd')
	data_inicio_heparina_cap = request.POST.get('data_inicio_heparina')

	if data_inicio_heparina_cap == '' or data_inicio_heparina_cap == None:
		if data_inicio_heparina_bd_cap == '' or data_inicio_heparina_bd_cap == None:
			data_inicio_heparina = None
		else:
			data_inicio_heparina = datetime.strptime(data_inicio_heparina_bd_cap, '%d/%m/%Y').date()
	else:
		data_inicio_heparina = data_inicio_heparina_cap
		

	tipo_heparina = request.POST.get('tipo_heparina')
	dose_heparina = request.POST.get('dose_heparina')
	
	pps_cap = request.POST.get('pps_paciente')
	if pps_cap == '' or pps_cap == None:
		pps = None
	else:
		pps = float(pps_cap)
		

	escala_pontos_glasgow_cap = request.POST.get('escala_glasgow')
	if escala_pontos_glasgow_cap == '' or escala_pontos_glasgow_cap == None:
		escala_pontos_glasgow = None
	else:
		escala_pontos_glasgow = int(float(escala_pontos_glasgow_cap))
		

	bloqueador_neuromuscular = request.POST.getlist('desc_bloqueador_neuromuscular')
	midazolam_dose = request.POST.get('midazolam_dose')
	fentanil_dose = request.POST.get('fentanil_dose')
	propofol_dose = request.POST.get('propofol_dose')
	sedacao_continua = request.POST.getlist('desc_sedacao_continua')
	rocuronio_dose = request.POST.get('rocuronio_dose')
	pacuronio_dose = request.POST.get('pacuronio_dose')
	cisatracurio_dose = request.POST.get('cisatracurio_dose')
	rass = request.POST.get('rass')
	
	data_exames_laboratorio_cap = request.POST.get('data_exames_laboratorio')
	data_exames_laboratorio_bd_cap = request.POST.get('data_exames_laboratorio_bd')

	if data_exames_laboratorio_cap == '' or data_exames_laboratorio_cap == None:
		if data_exames_laboratorio_bd_cap == '' or data_exames_laboratorio_bd_cap == None:
			data_exames_laboratorio = None
		else:
			data_exames_laboratorio = datetime.strptime(data_exames_laboratorio_bd_cap, '%d/%m/%Y').date()
	else:
		data_exames_laboratorio = data_exames_laboratorio_cap

	leucocito = request.POST.get('leucocito')
	linfocito = request.POST.get('linfocito')
	hb = request.POST.get('hb')
	ht = request.POST.get('ht')
	pcr = request.POST.get('pcr')
	lactato = request.POST.get('lactato')
	dhl = request.POST.get('dhl')
	ferritina = request.POST.get('ferritina')
	troponina = request.POST.get('troponina')
	cpk = request.POST.get('cpk')
	d_dimero = request.POST.get('d_dimero')
	ureia = request.POST.get('ureia')
	creatinina = request.POST.get('creatinina')
	ap = request.POST.get('ap')
	tap = request.POST.get('tap')
	inr = request.POST.get('inr')
	tgo = request.POST.get('tgo')
	tgp = request.POST.get('tgp')
	exame_imagem = request.POST.getlist('exame_imagem')
	laudo_tc_torax = request.POST.get('laudo_tc_torax')
	laudo_rx_torax = request.POST.get('laudo_rx_torax')
	is_sindrome_gripal = request.POST.get('sindrome_gripal')
	news_fast_pb = request.POST.get('new_fast')
	news_modificado = request.POST.get('new_modificado')
	uti = request.POST.get('uti')
	leito = request.POST.get('perfil')

	parecer_medico = request.POST.get('parecer')
	
	prioridade_cap = request.POST.get('prioridade')
	if prioridade_cap == '' or prioridade_cap == None:
		prioridade = None
	else:
		prioridade = int(float(prioridade_cap))


	descricao = request.POST.get('status_paciente')
	registro_covid = registro

	status = Status.objects.create(
		descricao=descricao,
		registro_covid=registro_covid
		)

	last_status = descricao

	if estabelecimento_referencia == '' or estabelecimento_referencia == None:
		estabelecimento_referencia_covid = estabelecimento_referencia_outro
	else:
		estabelecimento_referencia_covid = estabelecimento_referencia


	senha = 0
	if last_status == 'Regulado':
		#Regulações do paciente
		#regulacoes_registro = Regulacao.objects.filter(registro_covid_id=registro.id)
		#Ultima regulacao do paciente:
		codigo_sescovid = request.POST.get('num_protocolo')
		regulacao_last = Regulacao.objects.all().last()
		if regulacao_last:
			senha_last = regulacao_last.senha
			senha_new = senha_last + 1
		else:
			senha_new = senha + 500000
			

		senha = int(senha_new)

		regulacao = Regulacao.objects.create(
			estabelecimento_referencia_covid = estabelecimento_referencia_covid,
			nome_paciente = nome_paciente,
			senha = senha,
			registro_covid = registro_covid
			)


	justificativa = request.POST.get('justificativa')
	observacao = request.POST.get('observacoes_medicas')
	
	pareceristas = registro.pareceristas
	
	pareceristas_cap = request.POST.getlist('pareceristas')
	
	if pareceristas == '' or pareceristas == None:
		pareceristas = pareceristas_cap
	else:
		pareceristas = pareceristas + pareceristas_cap

	
	data_regulacao = datetime.now()
	#data_regulacao = data_regulacao_.strftime('%d/%m/%Y %H:%M')


	data_obito_bd_cap = request.POST.get('data_obito_bd')
	data_obito_cap = request.POST.get('data_obito')

	if data_obito_cap == '' or data_obito_cap == None:
		if data_obito_bd_cap == '' or data_obito_bd_cap == None:
			data_obito = None
		else:
			data_obito = datetime.strptime(data_obito_bd_cap, '%d/%m/%Y').date()
	else:
		data_obito = data_obito_cap

	
	registro.responsavel_pelo_preenchimento = responsavel_pelo_preenchimento
	#registro.codigo_registro = codigo_registro
	registro.nome_solicitante = nome_solicitante
	#registro.municipio_estabelecimento_solicitante = municipio_estabelecimento_solicitante
	#registro.estabelecimento_solicitante = estabelecimento_solicitante
	#registro.estabelecimento_solicitante_outro = estabelecimento_solicitante_outro
	registro.municipio_estabelecimento_referencia = municipio_estabelecimento_referencia
	registro.estabelecimento_referencia = estabelecimento_referencia
	registro.estabelecimento_referencia_outro = estabelecimento_referencia_outro
	registro.nome_paciente = nome_paciente
	registro.idade_paciente = idade_paciente
	registro.sexo_paciente = sexo_paciente
	registro.recurso_que_precisa = recurso_que_precisa
	registro.estado_origem = estado_origem
	registro.cidade_origem = cidade_origem
	registro.telefone_retorno = telefone_retorno
	registro.frequencia_respiratoria = frequencia_respiratoria
	registro.saturacao_paciente = saturacao_paciente
	registro.ar_o2 = ar_o2
	registro.frequencia_cardiaca = frequencia_cardiaca
	registro.pa = pa
	registro.conciencia = conciencia
	registro.temperatura = temperatura
	registro.descricao_clinica = descricao_clinica
	#registro.image_descricao_clinica = image_descricao_clinica
	registro.sindrome_gripal = sindrome_gripal
	registro.tempo_quadro_sintomatico = tempo_quadro_sintomatico
	registro.exposicao_pessoa_infectada = exposicao_pessoa_infectada
	registro.parentesco = parentesco
	registro.comorbidades = comorbidades
	registro.outras_comorbidades = outras_comorbidades
	registro.outras_cardiopatias = outras_cardiopatias
	registro.idade_gestacional = idade_gestacional
	registro.comorbidades_obstetricas = comorbidades_obstetricas
	registro.gesta_para = gesta_para
	registro.medicamentos_uso_habitual = medicamentos_uso_habitual
	registro.medicamento_outros = medicamento_outros
	registro.spo2 = spo2
	registro.fr_irpm = fr_irpm
	registro.ventilacao_tipo = ventilacao_tipo
	registro.o2_suporte = o2_suporte
	registro.dose_cn = dose_cn
	registro.dose_venturi = dose_venturi
	registro.vmi = vmi
	registro.vt = vt
	registro.delta_pressure = delta_pressure
	registro.pplato = pplato
	registro.raw = raw
	registro.cest = cest
	registro.sao2 = sao2
	registro.pao2 = pao2
	registro.fio2 = fio2
	registro.paco2 = paco2
	registro.pa = pa
	registro.fc = fc
	registro.temperatura_axilar = temperatura_axilar
	registro.droga_vasoativa = droga_vasoativa
	registro.qtd_nora = qtd_nora
	registro.qtd_adrenalina = qtd_adrenalina
	registro.qtd_vasopressina = qtd_vasopressina
	registro.qtd_dobutamina = qtd_dobutamina
	registro.arritmia = arritmia
	registro.infeccao_bacteriana = infeccao_bacteriana
	registro.foco = foco
	registro.uso_antibioticoterapia = uso_antibioticoterapia
	registro.pesquisa_teste_sars_cov_2 = pesquisa_teste_sars_cov_2
	registro.igg = igg
	registro.igm = igm
	registro.data_igm_igg = data_igm_igg
	registro.rt_pcr_sars_cov_2 = rt_pcr_sars_cov_2
	registro.data_coleta = data_coleta
	registro.em_uso_corticosteroide = em_uso_corticosteroide
	registro.dose_corticosteroide = dose_corticosteroide
	registro.data_inicio_corticosteroide = data_inicio_corticosteroide
	registro.em_uso_hidroxicloroquina = em_uso_hidroxicloroquina
	registro.data_inicio_hidroxicloroquina = data_inicio_hidroxicloroquina
	registro.em_uso_oseltamivir = em_uso_oseltamivir
	registro.data_inicio_oseltamivir = data_inicio_oseltamivir
	registro.em_uso_heparina = em_uso_heparina
	registro.data_inicio_heparina = data_inicio_heparina
	registro.tipo_heparina = tipo_heparina
	registro.dose_heparina = dose_heparina
	registro.pps = pps
	registro.escala_pontos_glasgow = escala_pontos_glasgow
	registro.bloqueador_neuromuscular = bloqueador_neuromuscular
	registro.midazolam_dose = midazolam_dose
	registro.fentanil_dose = fentanil_dose
	registro.propofol_dose = propofol_dose
	registro.sedacao_continua = sedacao_continua
	registro.rocuronio_dose = rocuronio_dose
	registro.pacuronio_dose = pacuronio_dose
	registro.cisatracurio_dose = cisatracurio_dose
	registro.rass = rass
	registro.data_exames_laboratorio = data_exames_laboratorio
	registro.leucocito = leucocito
	registro.linfocito = linfocito
	registro.hb = hb
	registro.ht = ht
	registro.pcr = pcr
	registro.lactato = lactato
	registro.dhl = dhl
	registro.ferritina = ferritina
	registro.troponina = troponina
	registro.cpk = cpk
	registro.d_dimero = d_dimero
	registro.ureia = ureia
	registro.creatinina = creatinina
	registro.ap = ap
	registro.tap = tap
	registro.inr = inr
	registro.tgo = tgo
	registro.tgp = tgp
	registro.exame_imagem = exame_imagem
	#registro.image_laudo_tc = image_laudo_tc
	#registro.image_laudo_rx = image_laudo_rx
	registro.laudo_tc_torax = laudo_tc_torax
	registro.laudo_rx_torax = laudo_rx_torax
	registro.is_sindrome_gripal = is_sindrome_gripal
	registro.news_fast_pb = news_fast_pb
	registro.news_modificado = news_modificado
	registro.uti = uti
	registro.leito = leito
	registro.parecer_medico = parecer_medico
	registro.prioridade = prioridade
	#registro.regulacao_paciente = regulacao_paciente
	#registro.status_regulacao = status_regulacao
	#registro.senha = senha
	registro.codigo_sescovid = codigo_sescovid
	registro.justificativa = justificativa
	registro.observacao = observacao
	registro.pareceristas = pareceristas
	registro.data_regulacao = data_regulacao
	registro.last_status = last_status
	registro.data_obito = data_obito

	registro.save()

	return redirect('regulacao_detail', id=id)


@login_required
def template_censo(request):
	n = RegistroCovid.objects.all()

	return render(request, 'template_censo.html', {'n':n})